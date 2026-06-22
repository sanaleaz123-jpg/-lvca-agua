"""
services/microcistina_import.py
Lectura del libro Excel "SOLVER MICROCISTINAS SAES" para importar una corrida
ELISA a la plataforma.

Por defecto TOMA LOS VALORES QUE YA CALCULÓ EL SOLVER del Excel (curva 4PL,
control y concentraciones por muestra) — no recalcula. Si el libro no trae los
valores calculados (curva vacía), cae a recalcular con el motor propio
(services/elisa_microcistina.py) y lo avisa.

Celdas (hoja "MCT SAES"):
    - Estándares OD: E11:J12 (cols 5..10 = Std0..Std5; filas 11/12 = réplicas).
    - Curva 4PL: A=E35, B=E36, C=E37, D=E38; R²=G35; SSE=G32.
    - Control: OD en D43/D44; concentración por pozo en G43/G44; promedio H44;
      %CV en F44.
    - Muestras: pares desde la fila 45 (paso 2). C=label, D=OD; en la 2.ª fila
      del par: F=%CV, I=concentración (mg/L).

Función pública:
    parse_excel_solver(origen) -> CorridaImportada
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO
from typing import Optional, Union

import openpyxl

from services.elisa_microcistina import (
    STD_CONC_UGL,
    FACTOR_DILUCION_DEFAULT,
    fit_4pl,
    procesar_muestra,
    procesar_control,
    concentracion_ugL,
    CurvaParams,
    ResultadoMuestra,
)

HOJA = "MCT SAES"

# ─────────────────────────────────────────────────────────────────────────────
# Distribución FIJA de la placa del LVCA (8 filas A–H × 12 columnas, duplicados
# en pares de columnas). Confirmada con el laboratorio:
#   Par 1 (col 1-2):  A..F = ST0..ST5, G = Control (CT), H = S1
#   Pares 2..6:       muestras "serpenteando" de abajo (H) hacia arriba (A)
#     col 3-4  → S2..S9     col 5-6  → S10..S17    col 7-8  → S18..S25
#     col 9-10 → S26..S33   col 11-12→ S34..S41
# Total: 6 estándares + control + 41 muestras.
# ─────────────────────────────────────────────────────────────────────────────


def _mapa_placa() -> dict:
    """
    Devuelve el mapeo de la distribución fija:
        {"std": [(row,col1,col2), ...x6], "control": (row,col1,col2),
         "samples": {n: (row,col1,col2)}}
    Filas A–H = 0..7; columnas 1..12 = 0..11.
    """
    std = [(i, 0, 1) for i in range(6)]          # ST0..ST5 en filas A–F, col 1-2
    control = (6, 0, 1)                           # CT en fila G, col 1-2
    samples = {1: (7, 0, 1)}                       # S1 en fila H, col 1-2
    for p in range(1, 6):                          # pares de columnas 3-4 .. 11-12
        c1, c2 = 2 * p, 2 * p + 1
        base = 2 + 8 * (p - 1)                     # 2, 10, 18, 26, 34
        for row in range(8):                       # A..H
            samples[base + (7 - row)] = (row, c1, c2)
    return {"std": std, "control": control, "samples": samples}


@dataclass
class MuestraImportada:
    label: str                       # "Sample 1", etc. (etiqueta del Excel)
    od_1: float
    od_2: float
    cv_pct: float
    conc_ugL: Optional[float]        # µg/L (con factor) — None si fuera de rango
    en_rango: bool
    motivo: str = ""


@dataclass
class CorridaImportada:
    curva: CurvaParams
    control: ResultadoMuestra
    control_conc_1: Optional[float]
    control_conc_2: Optional[float]
    muestras: list[MuestraImportada]
    std_od: list[tuple[float, float]] = field(default_factory=list)
    control_od: tuple[float, float] = (0.0, 0.0)
    kit_lote: Optional[str] = None
    orden: Optional[int] = None
    factor: float = FACTOR_DILUCION_DEFAULT
    recalculado: bool = False
    avisos: list[str] = field(default_factory=list)


def _num(ws, fila: int, col: int) -> Optional[float]:
    v = ws.cell(fila, col).value
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def parse_excel_solver(
    origen: Union[bytes, BytesIO, str],
    factor: float = FACTOR_DILUCION_DEFAULT,
    recalcular: bool = False,
) -> CorridaImportada:
    """
    Lee el libro del Solver. Por defecto usa los valores ya calculados por el
    Excel; con ``recalcular=True`` recalcula desde las OD con el motor propio.
    ``origen`` puede ser bytes, BytesIO o ruta.
    """
    if isinstance(origen, bytes):
        origen = BytesIO(origen)
    wb = openpyxl.load_workbook(origen, data_only=True)
    if HOJA not in wb.sheetnames:
        raise ValueError(
            f"El archivo no parece ser el SOLVER de microcistina: falta la hoja '{HOJA}'."
        )
    ws = wb[HOJA]
    avisos: list[str] = []

    # ── Estándares OD (E11:J12) — siempre se leen (para guardar/recalcular)
    std_od: list[tuple[float, float]] = []
    for i in range(len(STD_CONC_UGL)):
        col = 5 + i
        a = _num(ws, 11, col)
        b = _num(ws, 12, col)
        if a is None or b is None:
            raise ValueError(
                f"Faltan absorbancias del estándar {i} (celdas "
                f"{openpyxl.utils.get_column_letter(col)}11/12)."
            )
        std_od.append((a, b))

    # ── Curva 4PL: del Excel (E35:E38, G35, G32) o recalculada
    A = _num(ws, 35, 5); B = _num(ws, 36, 5)
    C = _num(ws, 37, 5); D = _num(ws, 38, 5)
    r2 = _num(ws, 35, 7); sse = _num(ws, 32, 7)
    if recalcular or None in (A, B, C, D):
        if not recalcular:
            avisos.append("El Excel no traía la curva calculada; se recalculó con el motor propio.")
        curva = fit_4pl(list(STD_CONC_UGL), [(a + b) / 2 for a, b in std_od])
        recalculado = True
    else:
        curva = CurvaParams(A=A, B=B, C=C, D=D, r2=r2 or 0.0, sse=sse or 0.0)
        recalculado = False
    if not curva.es_valida():
        avisos.append(
            f"La curva no cumple los criterios guía (A={curva.A:.3f}, "
            f"D={curva.D:.3f}, R²={curva.r2:.4f})."
        )

    # ── Metadatos
    kit_lote = ws.cell(7, 10).value  # J7
    kit_lote = str(kit_lote).strip() if kit_lote not in (None, "") else None
    orden = _num(ws, 13, 19)  # S13
    orden = int(orden) if orden is not None else None

    # ── Control: OD en D43/D44; conc por pozo G43/G44; prom H44; %CV F44
    ctrl_od1 = _num(ws, 43, 4)
    ctrl_od2 = _num(ws, 44, 4)
    if ctrl_od1 is None or ctrl_od2 is None:
        raise ValueError("Faltan las absorbancias del control (D43/D44).")
    if recalculado:
        control = procesar_control(ctrl_od1, ctrl_od2, curva)
        from services.elisa_microcistina import concentracion_ugL
        control_conc_1 = concentracion_ugL(ctrl_od1, curva)
        control_conc_2 = concentracion_ugL(ctrl_od2, curva)
    else:
        control_conc_1 = _num(ws, 43, 7)   # G43 (µg/L)
        control_conc_2 = _num(ws, 44, 7)   # G44 (µg/L)
        prom = _num(ws, 44, 8)             # H44 (µg/L)
        cv = _num(ws, 44, 6)               # F44 (%CV)
        control = ResultadoMuestra(
            od_1=ctrl_od1, od_2=ctrl_od2, cv_pct=cv or 0.0,
            conc_ugL=prom, en_rango=prom is not None,
        )

    # ── Muestras: pares desde la fila 45 (paso 2)
    muestras: list[MuestraImportada] = []
    r = 45
    while r <= ws.max_row:
        label = ws.cell(r, 3).value
        od1 = _num(ws, r, 4)
        od2 = _num(ws, r + 1, 4)
        if (label is None or str(label).strip() == "") and od1 is None:
            break
        if od1 is None or od2 is None:
            r += 2
            continue
        if recalculado:
            res = procesar_muestra(od1, od2, curva, factor=factor)
            cv_pct, conc_ugL, en_rango, motivo = (
                res.cv_pct, res.conc_ugL, res.en_rango, res.motivo
            )
        else:
            cv_pct = _num(ws, r + 1, 6) or 0.0        # F
            conc_mgL = _num(ws, r + 1, 9)             # I (mg/L)
            conc_ugL = conc_mgL * 1000.0 if conc_mgL is not None else None
            en_rango = conc_ugL is not None
            motivo = "" if en_rango else "Fuera del rango de la curva en el Excel."
        muestras.append(MuestraImportada(
            label=str(label).strip(),
            od_1=od1, od_2=od2, cv_pct=cv_pct,
            conc_ugL=conc_ugL, en_rango=en_rango, motivo=motivo,
        ))
        r += 2

    return CorridaImportada(
        curva=curva,
        control=control,
        control_conc_1=control_conc_1,
        control_conc_2=control_conc_2,
        muestras=muestras,
        std_od=std_od,
        control_od=(ctrl_od1, ctrl_od2),
        kit_lote=kit_lote,
        orden=orden,
        factor=factor,
        recalculado=recalculado,
        avisos=avisos,
    )


# ─────────────────────────────────────────────────────────────────────────────
# Placa OD cruda (8×12) pegada/subida tal cual sale del lector
# ─────────────────────────────────────────────────────────────────────────────

import re


def parse_grid_text(text: str) -> list[list[float]]:
    """
    Convierte el texto pegado de la placa en una matriz 8×12 de floats.
    Tolera etiquetas de fila (A–H), separadores por tab/espacio/coma/';' y
    coma decimal. Ignora una fila de encabezado de columnas (1..12) si aparece.
    """
    filas: list[list[float]] = []
    for linea in text.strip().splitlines():
        linea = linea.strip()
        if not linea:
            continue
        nums = re.findall(r"-?\d+(?:[.,]\d+)?", linea)
        vals = [float(n.replace(",", ".")) for n in nums]
        # Saltar fila de encabezado "1 2 3 ... 12"
        if vals == [float(i) for i in range(1, 13)]:
            continue
        if vals:
            filas.append(vals[:12] if len(vals) >= 12 else vals)
    if len(filas) != 8 or any(len(f) != 12 for f in filas):
        raise ValueError(
            "Se esperaban 8 filas (A–H) × 12 columnas de absorbancias. "
            f"Se leyeron {len(filas)} fila(s) con longitudes {[len(f) for f in filas]}."
        )
    return filas


def _grid_desde_hoja(ws) -> Optional[list[list[float]]]:
    """Busca en la hoja un bloque de 8 filas × 12 columnas de absorbancias.

    Tolera etiquetas de fila (A–H) y una fila de encabezado (1..12); toma las
    primeras 12 columnas numéricas de cada fila. Devuelve None si no encuentra
    8 filas válidas.
    """
    filas: list[list[float]] = []
    for row in ws.iter_rows(values_only=True):
        nums: list[float] = []
        for v in row:
            if isinstance(v, bool):
                continue
            if isinstance(v, (int, float)):
                nums.append(float(v))
            elif v is not None:
                try:
                    nums.append(float(str(v).strip().replace(",", ".")))
                except ValueError:
                    pass
        if len(nums) >= 12:
            cand = nums[:12]
            if cand == [float(i) for i in range(1, 13)]:
                continue  # fila de encabezado de columnas
            filas.append(cand)
            if len(filas) == 8:
                break
    return filas if len(filas) == 8 else None


def parse_placa_excel(
    origen: Union[bytes, BytesIO, str],
    factor: float = FACTOR_DILUCION_DEFAULT,
) -> CorridaImportada:
    """
    Lee un Excel que contiene SOLO la placa de absorbancias (8×12) — tal como
    sale del lector — y calcula la corrida con la distribución fija del LVCA.
    Busca el bloque 8×12 en cualquier hoja del libro.
    """
    if isinstance(origen, bytes):
        origen = BytesIO(origen)
    wb = openpyxl.load_workbook(origen, data_only=True)
    grid = None
    for nombre in wb.sheetnames:
        grid = _grid_desde_hoja(wb[nombre])
        if grid is not None:
            break
    if grid is None:
        raise ValueError(
            "No se encontró una placa de 8 filas × 12 columnas de absorbancias "
            "en el archivo. Verifica que el Excel contenga la placa completa."
        )
    return parse_placa_cruda(grid, factor=factor)


def parse_placa_cruda(
    grid: list[list[float]],
    factor: float = FACTOR_DILUCION_DEFAULT,
) -> CorridaImportada:
    """
    Calcula la corrida a partir de la placa de OD cruda (8×12) usando la
    distribución fija del LVCA y el motor propio (curva 4PL + concentración).
    """
    if len(grid) != 8 or any(len(r) != 12 for r in grid):
        raise ValueError("La placa debe ser 8 filas × 12 columnas.")
    g = [[float(v) for v in row] for row in grid]
    mapa = _mapa_placa()
    avisos: list[str] = []

    # Estándares y curva
    std_od = [(g[r][c1], g[r][c2]) for (r, c1, c2) in mapa["std"]]
    curva = fit_4pl(list(STD_CONC_UGL), [(a + b) / 2 for a, b in std_od])
    if not curva.es_valida():
        avisos.append(
            f"La curva no cumple los criterios guía (A={curva.A:.3f}, "
            f"D={curva.D:.3f}, R²={curva.r2:.4f})."
        )

    # Control
    cr, cc1, cc2 = mapa["control"]
    ctrl_od1, ctrl_od2 = g[cr][cc1], g[cr][cc2]
    control = procesar_control(ctrl_od1, ctrl_od2, curva)
    control_conc_1 = concentracion_ugL(ctrl_od1, curva)
    control_conc_2 = concentracion_ugL(ctrl_od2, curva)

    # Muestras (S1..S41) en orden
    muestras: list[MuestraImportada] = []
    for n in sorted(mapa["samples"].keys()):
        r, c1, c2 = mapa["samples"][n]
        od1, od2 = g[r][c1], g[r][c2]
        res = procesar_muestra(od1, od2, curva, factor=factor)
        muestras.append(MuestraImportada(
            label=f"S{n}", od_1=od1, od_2=od2, cv_pct=res.cv_pct,
            conc_ugL=res.conc_ugL, en_rango=res.en_rango, motivo=res.motivo,
        ))

    return CorridaImportada(
        curva=curva, control=control,
        control_conc_1=control_conc_1, control_conc_2=control_conc_2,
        muestras=muestras, std_od=std_od, control_od=(ctrl_od1, ctrl_od2),
        kit_lote=None, orden=2, factor=factor, recalculado=True, avisos=avisos,
    )
