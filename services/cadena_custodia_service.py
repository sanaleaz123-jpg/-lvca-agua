"""
services/cadena_custodia_service.py
Generación de Cadena de Custodia en formato Excel y PDF.
Formato oficial AUTODEMA — CC-MON-01 Rev 03.

El Excel se genera a partir de la plantilla original (templates/cadena_template.xlsx)
preservando exactamente el formato, colores, fuentes, merges y estructura.

Funciones públicas:
    get_datos_cadena(campana_id)                 → datos para la cadena
    generar_excel_cadena(campana_id, config)      → bytes .xlsx
    generar_pdf_cadena(campana_id, config)        → bytes .pdf
"""

from __future__ import annotations

import os
from datetime import datetime
from io import BytesIO
from typing import Optional

from database.client import get_admin_client
from services.parametro_registry import (
    get_parametros_lab_cadena,
    get_parametros_campo_cadena,
    get_insitu_a_cadena_map,
    get_all_param_configs,
)


# ─────────────────────────────────────────────────────────────────────────────
# Helpers compartidos por Excel y PDF — selección dinámica de parámetros
# ─────────────────────────────────────────────────────────────────────────────

def _params_lab_seleccionados(cfg: dict, max_cols: int) -> list[dict]:
    """
    Filtra los parámetros de laboratorio a incluir como columnas del
    documento, basándose en la selección guardada en la campaña.

    Si la campaña no tiene selección guardada, devuelve todos los lab
    disponibles (hasta ``max_cols``). Los parámetros adicionales escritos
    por el usuario (``parametros_lab_extra``) se añaden al final como
    entradas sin código — sin preservante ni tipo de frasco asociado.
    """
    all_lab = get_parametros_lab_cadena()
    claves_seleccionadas = cfg.get("parametros_lab")
    if claves_seleccionadas:
        sel = set(claves_seleccionadas)
        filtrados = [p for p in all_lab if p["clave"] in sel]
    else:
        filtrados = list(all_lab)

    extras = cfg.get("parametros_lab_extra") or []
    for extra in extras:
        nombre = str(extra).strip()
        if not nombre:
            continue
        filtrados.append({
            "clave":  nombre.lower().replace(" ", "_"),
            "nombre": nombre,
            "codigo": "",
        })
    return filtrados[:max_cols]


def _preservante_de_param(codigo: str, param_configs: dict) -> str:
    """
    Retorna el preservante configurado para un parámetro, o "S/P" si no tiene.
    Normaliza: cualquier valor "Ninguno" o vacío se mapea a "S/P".
    """
    pres = (param_configs.get(codigo) or {}).get("preservante", "")
    if not pres or pres == "Ninguno":
        return "S/P"
    return pres


def _contar_botellas(lab_params: list[dict], param_configs: dict) -> tuple[int, int]:
    """
    Cuenta botellas V (vidrio) y P (plástico) necesarias para la muestra.

    Agrupa por (preservante, tipo_frasco) para no contar múltiples parámetros
    que comparten botella. Clasifica cada grupo:
        - V: si el tipo_frasco contiene "Vidrio"
        - P: el resto (Polipropileno u otro plástico)
    """
    grupos: set[tuple[str, str]] = set()
    for p in lab_params:
        codigo = p.get("codigo", "")
        if not codigo:
            continue  # parámetros extra no tienen botella definida
        cfg_p = param_configs.get(codigo) or {}
        tf = (cfg_p.get("tipo_frasco") or "").strip()
        if not tf:
            continue
        pres = cfg_p.get("preservante") or "Ninguno"
        grupos.add((pres, tf))
    n_vidrio = sum(1 for _, tf in grupos if "vidrio" in tf.lower())
    n_plastico = sum(1 for _, tf in grupos if "vidrio" not in tf.lower())
    return n_vidrio, n_plastico


# ─────────────────────────────────────────────────────────────────────────────
# Parámetros de laboratorio y campo — ahora dinámicos desde la BD
# Se mantienen los nombres para retrocompatibilidad con importaciones.
# ─────────────────────────────────────────────────────────────────────────────

def _get_parametros_lab_default() -> list[dict]:
    """Parámetros de laboratorio para la cadena, leídos de la BD."""
    return get_parametros_lab_cadena()

def _get_parametros_campo() -> list[dict]:
    """Parámetros de campo para la cadena, leídos de la BD."""
    return get_parametros_campo_cadena()

# Accesores globales para retrocompatibilidad de importación
PARAMETROS_LAB_DEFAULT = _get_parametros_lab_default
PARAMETROS_CAMPO = _get_parametros_campo

EQUIPOS_DEFAULT: list[dict] = [
    {"codigo": "21G102303/N",  "nombre": "Equipo multiparametro PRODSS YSI"},
    {"codigo": "9208180023",   "nombre": "Turbidimetro PALINTEST"},
]


def get_equipos_registrados() -> list[dict]:
    """
    Retorna la lista de equipos registrados en la tabla 'equipos_medicion'.
    Si la tabla no existe, retorna EQUIPOS_DEFAULT.
    """
    try:
        db = get_admin_client()
        res = (
            db.table("equipos_medicion")
            .select("id, codigo, nombre, activo")
            .eq("activo", True)
            .order("nombre")
            .execute()
        )
        equipos = res.data or []
        if equipos:
            return equipos
    except Exception:
        pass
    # Fallback a la lista por defecto
    return EQUIPOS_DEFAULT


def registrar_equipo(codigo: str, nombre: str) -> dict:
    """Registra un nuevo equipo de medición."""
    try:
        db = get_admin_client()
        res = db.table("equipos_medicion").insert({
            "codigo": codigo,
            "nombre": nombre,
            "activo": True,
        }).execute()
        return res.data[0]
    except Exception:
        # Si la tabla no existe, solo retornar el dict
        return {"codigo": codigo, "nombre": nombre}

# Mapeo de claves insitu a claves de la cadena — función para evaluación dinámica
def _get_insitu_map():
    return get_insitu_a_cadena_map()

# Mapeo dinámico: preservante → claves de parámetros lab que lo requieren.
# Se construye desde la configuración en data/parametros_config.json
# para que cualquier cambio en el módulo Parámetros se propague automáticamente.
def _get_preservante_claves() -> dict[str, set[str]]:
    """Construye {preservante: {codigo_lower, ...}} desde la config de parámetros."""
    from services.parametro_registry import get_all_param_configs
    mapping: dict[str, set[str]] = {}
    for codigo, cfg in get_all_param_configs().items():
        if not isinstance(cfg, dict):
            continue
        pres = cfg.get("preservante", "Ninguno")
        if pres and pres != "Ninguno":
            mapping.setdefault(pres, set()).add(codigo.lower())
    return mapping

# Fila en el template donde va cada preservante
_PRESERVANTE_ROW = {
    "HNO3": 4,
    "H2SO4": 5,
    "HCl": 6,
    "Lugol": 7,
    "S/P": 8,
    "Formol": 7,  # Comparte fila con Lugol (biológicos)
}

# Columna base de lab params en el template (15 columnas originales)
_COL_LAB_START = 26   # Z
_TEMPLATE_LAB_COUNT = 15
_TEMPLATE_FIELD_COUNT = 7

# Ruta de la plantilla
_TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "..", "templates", "cadena_template.xlsx"
)


# ─────────────────────────────────────────────────────────────────────────────
# Obtener datos para la cadena
# ─────────────────────────────────────────────────────────────────────────────

def get_datos_cadena(campana_id: str) -> dict:
    """
    Recopila todos los datos necesarios para generar la cadena de custodia:
        campana, muestras (con punto, coordenadas, mediciones insitu)
    """
    db = get_admin_client()

    # Campaña
    campana = (
        db.table("campanas")
        .select(
            "id, codigo, nombre, fecha_inicio, fecha_fin, estado, "
            "responsable_campo, responsable_laboratorio, observaciones"
        )
        .eq("id", campana_id)
        .single()
        .execute()
        .data
    )

    # Muestras con punto (coordenadas UTM, cuenca, etc.)
    _select_muestras = (
        "id, codigo, fecha_muestreo, hora_recoleccion, tipo_muestra, "
        "estado, preservante, observaciones_campo, "
        "clima, caudal_estimado, nivel_agua, temperatura_transporte, "
        "puntos_muestreo(codigo, nombre, tipo, cuenca, sistema_hidrico, "
        "  utm_este, utm_norte, utm_zona, altitud_msnm, "
        "  latitud, longitud)"
    )
    _depth_fields = (
        ", modo_muestreo, profundidad_tipo, profundidad_valor, "
        "grupo_profundidad, profundidad_total, profundidad_secchi"
    )
    try:
        db.table("muestras").select("modo_muestreo").limit(1).execute()
        _select_muestras += _depth_fields
    except Exception:
        pass
    m_res = (
        db.table("muestras")
        .select(_select_muestras)
        .eq("campana_id", campana_id)
        .order("fecha_muestreo")
        .execute()
    )
    muestras = m_res.data or []

    # Mediciones in situ por muestra
    muestra_ids = [m["id"] for m in muestras]
    insitu_map: dict[str, dict] = {}

    if muestra_ids:
        i_res = (
            db.table("mediciones_insitu")
            .select("muestra_id, parametro, valor")
            .in_("muestra_id", muestra_ids)
            .execute()
        )
        insitu_key_map = _get_insitu_map()
        for r in (i_res.data or []):
            mid = r["muestra_id"]
            if mid not in insitu_map:
                insitu_map[mid] = {}
            clave_cadena = insitu_key_map.get(r["parametro"], r["parametro"])
            insitu_map[mid][clave_cadena] = r["valor"]

    # Enriquecer cada muestra con insitu
    for m in muestras:
        m["insitu"] = insitu_map.get(m["id"], {})

    # Inferir profundidad_tipo si no viene de la BD (migración 005 pendiente)
    # Agrupar por punto: si hay 3 muestras del mismo punto, asignar S/M/F
    _prof_labels = {0: "S", 1: "M", 2: "F"}
    _prof_nombres = {"S": "Superficie", "M": "Medio", "F": "Fondo"}
    muestras_por_punto: dict[str, list[dict]] = {}
    for m in muestras:
        pt_id = (m.get("puntos_muestreo") or {}).get("codigo", "")
        muestras_por_punto.setdefault(pt_id, []).append(m)
    for pt_id, ms in muestras_por_punto.items():
        if len(ms) >= 3:
            for i, m in enumerate(ms):
                if not m.get("profundidad_tipo") and i < 3:
                    m["profundidad_tipo"] = _prof_labels[i]
                # Agregar label legible para observaciones
                tp = m.get("profundidad_tipo")
                if tp and tp in _prof_nombres:
                    m["_prof_label"] = _prof_nombres[tp]

    # Lugar de monitoreo (del primer punto)
    lugar = ""
    cuenca = ""
    if muestras and muestras[0].get("puntos_muestreo"):
        pt0 = muestras[0]["puntos_muestreo"]
        lugar = pt0.get("nombre", "")
        cuenca = pt0.get("cuenca", "")

    return {
        "campana": campana,
        "muestras": muestras,
        "lugar": lugar,
        "cuenca": cuenca,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Configuración por defecto
# ─────────────────────────────────────────────────────────────────────────────

def get_config_persistida(campana_id: str) -> dict | None:
    """
    Recupera la configuración guardada de cadena de custodia para una campaña.
    Retorna None si nunca se guardó.
    Requiere migración 006 (tabla cadena_custodia_config).
    """
    try:
        from database.client import get_admin_client
        db = get_admin_client()
        res = (
            db.table("cadena_custodia_config")
            .select("config")
            .eq("campana_id", campana_id)
            .maybe_single()
            .execute()
        )
        if res and res.data:
            return res.data.get("config")
    except Exception:
        pass
    return None


def guardar_config_persistida(
    campana_id: str,
    config: dict,
    usuario_id: str | None = None,
) -> bool:
    """
    Persiste la configuración de cadena para reutilizarla en próximas
    generaciones. Upsert por campana_id.
    Retorna True si se guardó, False si la tabla no existe (migración 006 pendiente).
    """
    try:
        from database.client import get_admin_client
        db = get_admin_client()
        payload = {
            "campana_id":      campana_id,
            "config":          config,
            "actualizado_por": usuario_id,
            "updated_at":      datetime.utcnow().isoformat(),
        }
        db.table("cadena_custodia_config").upsert(
            payload, on_conflict="campana_id"
        ).execute()
        return True
    except Exception:
        return False


def config_para_campana(campana_id: str) -> dict:
    """
    Devuelve la configuración a usar para una campaña:
        - persistida si existe
        - default si no
    """
    persistida = get_config_persistida(campana_id)
    if persistida:
        # Mezclar con defaults para llenar campos faltantes (compat futura)
        base = config_default()
        base.update(persistida)
        return base
    return config_default()


def config_default() -> dict:
    """Retorna la configuración por defecto de la cadena."""
    return {
        "codigo_documento": "CC-MON-01",
        "revision": "03",
        "area": "SUBGERENCIA DE OPERACIÓN Y MANTENIMIENTO - LABORATORIO DE VIGILANCIA DE CALIDAD DE AGUA",
        "direccion": "Urb. La Marina E-8 - Cayma - Arequipa",
        "telefono": "(054)254040",
        "institucion": "AUTORIDAD AUTONOMA DE MAJES",
        "urgencia": "Regular",
        "preservacion": {"HNO3": True, "H2SO4": True, "HCl": False, "Lugol": True, "Formol": False, "S/P": True},
        "condiciones": {"refrigerado": True, "icepack": True, "temp_ambiente": False,
                        "congelado": False, "caja_conservadora": False, "hielo_potable": False},
        "parametros_lab": [p["clave"] for p in _get_parametros_lab_default()],
        "parametros_lab_extra": [],
        "equipos": EQUIPOS_DEFAULT.copy(),
        "muestreo_por": "laboratorio",
        "nombre_muestreador": "",
        "nombre_receptor": "",
        "nombre_supervisor": "",
        "observaciones_generales": "",
    }


# ─────────────────────────────────────────────────────────────────────────────
# Helper: escritura segura en celdas (maneja MergedCells)
# ─────────────────────────────────────────────────────────────────────────────

def _safe_set(ws, row: int, col: int, value):
    """Escribe un valor en una celda; si es MergedCell, lo ignora."""
    try:
        ws.cell(row=row, column=col).value = value
    except (AttributeError, TypeError):
        pass


# ─────────────────────────────────────────────────────────────────────────────
# Generación Excel — basada en plantilla original
# ─────────────────────────────────────────────────────────────────────────────

_SLOTS_POR_HOJA = 11


def generar_excel_cadena(campana_id: str, config: dict | None = None) -> bytes:
    """
    Genera el Excel de cadena de custodia a partir de la plantilla AUTODEMA.

    Carga templates/cadena_template.xlsx, actualiza campos dinámicos y datos
    de las muestras, y devuelve los bytes del .xlsx resultante.

    La plantilla tiene estructura fija: 15 columnas lab (Z–AN) y 7 campo
    (AO–AU).  Los nombres de los headers se sobreescriben dinámicamente
    con los parámetros activos de la BD.

    Si la campaña tiene más de 11 muestras, genera hojas adicionales
    (Página 1, Página 2, ...) preservando el formato del template.
    """
    from openpyxl import load_workbook

    cfg = config or config_default()
    datos = get_datos_cadena(campana_id)
    campana = datos["campana"]
    muestras = datos["muestras"]
    n_muestras = len(muestras)
    n_hojas = max(1, (n_muestras + _SLOTS_POR_HOJA - 1) // _SLOTS_POR_HOJA)
    chunks = [
        muestras[i : i + _SLOTS_POR_HOJA]
        for i in range(0, max(n_muestras, 1), _SLOTS_POR_HOJA)
    ] or [[]]

    # Parámetros dinámicos — solo los seleccionados en la campaña aparecen
    # como columnas. El template tiene 15 slots máximos de lab y 7 de campo.
    all_campo = _get_parametros_campo()
    lab_params = _params_lab_seleccionados(cfg, _TEMPLATE_LAB_COUNT)
    campo_params = all_campo[:_TEMPLATE_FIELD_COUNT]

    # Configuración preservante/tipo_frasco por código de parámetro
    param_configs = get_all_param_configs()

    # Posiciones fijas del template — NO se insertan columnas
    col_field_start = _COL_LAB_START + _TEMPLATE_LAB_COUNT   # 41 (AO)
    col_obs = col_field_start + _TEMPLATE_FIELD_COUNT          # 48 (AV)

    # ── Cargar plantilla ──────────────────────────────────────────────────
    wb = load_workbook(_TEMPLATE_PATH)
    ws = wb.active
    _ws_actual = [ws]  # contenedor mutable para que _set use la hoja actual
    _set = lambda r, c, v: _safe_set(_ws_actual[0], r, c, v)

    # Si hay más de 11 muestras, duplicamos la hoja una vez por cada chunk
    hojas: list = [ws]
    for i in range(1, n_hojas):
        nueva = wb.copy_worksheet(ws)
        nueva.title = f"Pagina {i + 1}"
        hojas.append(nueva)
    if n_hojas > 1:
        ws.title = "Pagina 1"

    # ── Sobreescribir headers de parámetros (fila 11) ────────────────────
    for i in range(_TEMPLATE_LAB_COUNT):
        if i < len(lab_params):
            _set(11, _COL_LAB_START + i, lab_params[i]["nombre"])
        else:
            _set(11, _COL_LAB_START + i, "")

    for i in range(_TEMPLATE_FIELD_COUNT):
        if i < len(campo_params):
            _set(11, col_field_start + i, campo_params[i]["nombre"])
        else:
            _set(11, col_field_start + i, "")

    # ══════════════════════════════════════════════════════════════════════
    # 1. ENCABEZADO (filas 2–12)
    # ══════════════════════════════════════════════════════════════════════

    # Revisión documento (AO4 = col 41, merge AO4:AW8 — posición fija)
    _set(4, 41,
         f"Revision documento:\n"
         f"Codigo: {cfg.get('codigo_documento', 'CC-MON-01')}\n"
         f"Revision : {cfg.get('revision', '03')}\n"
         f"Fecha: {datetime.utcnow().strftime('%d-%b-%Y')}")

    # Área / Sub-área (F4 = col 6, merge F4:W5)
    _set(4, 6, cfg.get("area", ""))

    # Dirección (F6 = col 6, merge F6:J7)
    _set(6, 6, cfg.get("direccion", "Urb. La Marina E-8 - Cayma - Arequipa"))

    # Teléfono (O6 = col 15, merge O6:Q7)
    _set(6, 15, cfg.get("telefono", "(054)254040"))

    # Urgencia (T6 = col 20, merge T6:W6 / T7:W7)
    urgencia = cfg.get("urgencia", "Regular")
    _set(6, 20, f"Regular ({'X' if urgencia == 'Regular' else '    '})")
    _set(7, 20, f"Alta ({'X' if urgencia == 'Alta' else '    '})")

    # Institución (F8 = col 6, merge F8:W9)
    _set(8, 6, cfg.get("institucion", "AUTORIDAD AUTONOMA DE MAJES"))

    # Campaña (antes "Lugar de monitoreo") (F10 = col 6, merge F10:W11)
    campana_label = cfg.get("campana_label") or datos.get("lugar", "")
    _set(10, 6, campana_label)

    # Nombre de cuenca (F12 = col 6, merge F12:W12)
    _set(12, 6, datos.get("cuenca", ""))

    # ══════════════════════════════════════════════════════════════════════
    # 2. MARCAS DE PRESERVANTES (filas 4–8, cols 26–40)
    # ══════════════════════════════════════════════════════════════════════

    # Limpiar marcas existentes en las 15 columnas lab
    for row in range(4, 9):
        for col in range(_COL_LAB_START, _COL_LAB_START + _TEMPLATE_LAB_COUNT):
            _set(row, col, None)

    # Colocar marcas de preservante en cada columna seleccionada, tomando el
    # preservante desde la configuración del parámetro (data/parametros_config.json).
    for idx, p in enumerate(lab_params):
        codigo = p.get("codigo", "")
        preservante_asignado = _preservante_de_param(codigo, param_configs)
        row_pres = _PRESERVANTE_ROW.get(preservante_asignado)
        if row_pres:
            _set(row_pres, _COL_LAB_START + idx, "X")

    # ══════════════════════════════════════════════════════════════════════
    # 3. DATOS DE MUESTRAS (filas 15–36: 11 slots × 2 filas) — multi-hoja
    # ══════════════════════════════════════════════════════════════════════

    # Conteo de botellas: igual para todas las muestras de la campaña.
    n_vidrio, n_plastico = _contar_botellas(lab_params, param_configs)

    for hoja_idx, muestras_chunk in enumerate(chunks):
      _ws_actual[0] = hojas[hoja_idx]
      base_offset = hoja_idx * _SLOTS_POR_HOJA

      for slot in range(_SLOTS_POR_HOJA):
        r1 = 15 + slot * 2   # fila impar (E: / alt:)
        r2 = r1 + 1          # fila par   (N: / Z:)

        # ── Limpiar celdas del slot ───────────────────────────────────────
        for col in [2, 3, 6, 17, 18, 19, 24, 25, col_obs]:
            _set(r1, col, None)
        for col in [20, 21, 22, 23]:
            _set(r1, col, None)
            _set(r2, col, None)
        for col in range(_COL_LAB_START, _COL_LAB_START + _TEMPLATE_LAB_COUNT):
            _set(r1, col, None)
        for col in range(col_field_start, col_field_start + _TEMPLATE_FIELD_COUNT):
            _set(r1, col, None)

        # ── Si no hay muestra para este slot, saltar ─────────────────────
        if slot >= len(muestras_chunk):
            continue

        m = muestras_chunk[slot]
        pt = m.get("puntos_muestreo") or {}
        insitu = m.get("insitu", {})

        # N° global (B = col 2, merge B:B r1:r2)
        _set(r1, 2, float(base_offset + slot + 1))

        # Código de laboratorio (C = col 3, merge C:E r1:r2)
        _set(r1, 3, m.get("codigo", ""))

        # Punto de muestreo (F = col 6, merge F:P r1:r2)
        # Agregar sufijo de profundidad si es muestra de columna
        nombre_punto = pt.get("nombre", "")
        codigo_punto = pt.get("codigo", "")
        prof_tipo = m.get("profundidad_tipo")
        prof_sufijo_map = {"S": " (S)", "M": " (M)", "F": " (F)"}
        prof_sufijo = prof_sufijo_map.get(prof_tipo, "")
        if nombre_punto:
            _set(r1, 6, f"{nombre_punto} /{codigo_punto}{prof_sufijo}")
        else:
            _set(r1, 6, "")

        # Fecha (Q = col 17, merge Q:Q r1:r2)
        fecha_raw = m.get("fecha_muestreo", "")
        try:
            fecha_dt = datetime.strptime(str(fecha_raw)[:10], "%Y-%m-%d")
            _set(r1, 17, fecha_dt)
        except (ValueError, TypeError):
            _set(r1, 17, str(fecha_raw)[:10] if fecha_raw else None)

        # Hora (R = col 18, merge R:R r1:r2)
        _set(r1, 18, m.get("hora_recoleccion", "") or "")

        # Tipo de muestra / Matriz (S = col 19, merge S:S r1:r2)
        tipo_mapa = {
            "laguna": "ADL", "rio": "ADR", "canal": "ADR",
            "manantial": "AMA", "embalse": "ADL", "pozo": "ASUB",
        }
        _set(r1, 19, tipo_mapa.get(pt.get("tipo", ""), "AN"))

        # Coordenadas UTM — NO merged entre r1/r2 (valores distintos)
        _set(r1, 20, "E:")
        _set(r1, 21, pt.get("utm_este"))
        _set(r1, 22, "alt: ")
        _set(r1, 23, pt.get("altitud_msnm"))

        _set(r2, 20, "N:")
        _set(r2, 21, pt.get("utm_norte"))
        _set(r2, 22, "Z: ")
        _set(r2, 23, pt.get("utm_zona", "19 L"))

        # N° de frascos (X=24 vidrio, Y=25 plástico) — derivado del
        # tipo_frasco de cada parámetro seleccionado (constante en la cadena).
        _set(r1, 24, float(n_vidrio))
        _set(r1, 25, float(n_plastico))

        # Parámetros de laboratorio — "x" en todas las columnas (todas están
        # seleccionadas por construcción de lab_params).
        for i, _p in enumerate(lab_params):
            _set(r1, _COL_LAB_START + i, "x")

        # Parámetros de campo — valores numéricos con coma decimal (máx 7)
        for i, p in enumerate(campo_params):
            val = insitu.get(p["clave"])
            if val is not None:
                _set(r1, col_field_start + i, str(val).replace(".", ","))

        # Observaciones (AV = col 48, posición fija del template)
        # Auto-incluir clima, nivel, descarga, profundidad y temp. transporte
        obs_parts = []
        if m.get("clima"):
            obs_parts.append(f"Clima: {m['clima']}")
        if m.get("caudal_estimado"):
            obs_parts.append(f"Descarga: {m['caudal_estimado']}")
        if m.get("nivel_agua"):
            obs_parts.append(f"Nivel: {m['nivel_agua']}")
        if m.get("temperatura_transporte") is not None:
            obs_parts.append(f"T.transp: {m['temperatura_transporte']}°C")
        # Datos de profundidad
        if m.get("profundidad_total") is not None:
            obs_parts.append(f"Prof.total: {m['profundidad_total']}m")
        if m.get("profundidad_secchi") is not None:
            obs_parts.append(f"Secchi: {m['profundidad_secchi']}m")
        if m.get("profundidad_valor") is not None and m.get("_prof_label"):
            obs_parts.append(f"Prof: {m['_prof_label']} ({m['profundidad_valor']}m)")
        elif m.get("profundidad_valor") is not None:
            obs_parts.append(f"Prof.muestra: {m['profundidad_valor']}m")
        elif m.get("_prof_label"):
            obs_parts.append(f"Prof: {m['_prof_label']}")
        if m.get("observaciones_campo"):
            obs_parts.append(m["observaciones_campo"])
        obs = "; ".join(obs_parts)
        if obs:
            _set(r1, col_obs, obs)

    # ══════════════════════════════════════════════════════════════════════
    # 4. PIE DE PÁGINA (filas 38–65) — replicado en cada hoja
    # ══════════════════════════════════════════════════════════════════════

    fecha_hoy = datetime.utcnow().strftime("%d/%m/%Y")
    fecha_corta = datetime.utcnow().strftime("%d/%m/%y")
    fecha_hora = datetime.utcnow().strftime("%d/%m/%y - %H:%M")
    muestreo = cfg.get("muestreo_por", "laboratorio")
    equipos = cfg.get("equipos", EQUIPOS_DEFAULT)
    cond = cfg.get("condiciones", {})

    def _footer(set_fn):
        set_fn(39, 4, "(   X      )" if muestreo == "laboratorio" else "(         )")
        set_fn(39, 8, "(   X      )" if muestreo != "laboratorio" else "(         )")
        set_fn(40, 2, f"Nombre:  {cfg.get('nombre_muestreador', '')}")
        set_fn(40, 5, f"Nombre: {cfg.get('nombre_receptor', '')}")
        set_fn(42, 2, f"Fecha: {fecha_hoy}")
        set_fn(42, 5, "Fecha:")
        set_fn(45, 2, "Firma:")
        set_fn(45, 5, "Firma:")
        if len(equipos) > 0:
            set_fn(40, 18, equipos[0].get("codigo", ""))
            set_fn(40, 23, equipos[0].get("nombre", ""))
        else:
            set_fn(40, 18, "")
            set_fn(40, 23, "")
        if len(equipos) > 1:
            set_fn(44, 18, equipos[1].get("codigo", ""))
            set_fn(44, 23, equipos[1].get("nombre", ""))
        else:
            set_fn(44, 18, "")
            set_fn(44, 23, "")
        set_fn(50, 2, f"Nombre: {cfg.get('nombre_supervisor', '')}")
        set_fn(52, 2, f"Fecha: {fecha_corta}")
        set_fn(54, 2, "Firma:")
        set_fn(51, 10, f"Nombre: {cfg.get('nombre_receptor', '')}")
        set_fn(54, 10, f"Fecha / Hora: {fecha_hora}")
        set_fn(39, 49, f"({'X' if cond.get('temp_ambiente') else '   '})")
        set_fn(42, 49, f"({'x' if cond.get('refrigerado') else '    '})")
        set_fn(47, 49, f"({'x' if cond.get('congelado') else '    '})")
        set_fn(50, 49, f"({'x' if cond.get('caja_conservadora') else '    '})")
        set_fn(53, 49, f"({'x' if cond.get('icepack') else '    '})")
        set_fn(55, 49, f"({'x' if cond.get('hielo_potable') else '    '})")
        set_fn(60, 2, cfg.get("observaciones_generales", ""))

    for _h in hojas:
        _ws_actual[0] = _h
        _footer(_set)

    # ══════════════════════════════════════════════════════════════════════
    # 5. GUARDAR
    # ══════════════════════════════════════════════════════════════════════

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Generación PDF
# ─────────────────────────────────────────────────────────────────────────────

def generar_pdf_cadena(campana_id: str, config: dict | None = None) -> bytes:
    """
    Genera el PDF de cadena de custodia formato AUTODEMA landscape A4.
    Usa canvas de reportlab para control total: texto rotado, colores,
    estructura fiel al template Excel original.
    """
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfgen.canvas import Canvas
    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor, black, white, Color

    cfg = config or config_default()
    datos = get_datos_cadena(campana_id)
    campana = datos["campana"]
    muestras = datos["muestras"]

    # Solo los parámetros seleccionados en la campaña son columnas del PDF.
    params_lab = _params_lab_seleccionados(cfg, _TEMPLATE_LAB_COUNT)
    param_configs = get_all_param_configs()

    # Conteo de botellas (V/P) — se calcula una vez, común a toda la cadena.
    n_vidrio, n_plastico = _contar_botellas(params_lab, param_configs)

    # Preservantes activos según selección (para el bloque PRESERVACIÓN del header)
    preservantes_usados = set(
        _preservante_de_param(p.get("codigo", ""), param_configs)
        for p in params_lab
        if p.get("codigo")
    )

    output = BytesIO()
    W, H = landscape(A4)
    c = Canvas(output, pagesize=(W, H))

    # Colores del template AUTODEMA
    GREEN = HexColor("#A8D08D")
    LIGHT_GREEN = HexColor("#E2EFD9")
    GRAY_BG = HexColor("#F5F5F5")
    BORDER = HexColor("#808080")

    ML = 8 * mm   # margen izquierdo
    MR = 8 * mm   # margen derecho
    CW = W - ML - MR  # ancho útil

    # ── Helpers de dibujo ─────────────────────────────────────────────────

    def rect(x, y, w, h, fill=None):
        """Dibuja un rectángulo con borde y opcionalmente relleno."""
        if fill:
            c.setFillColor(fill)
            c.rect(x, y, w, h, fill=1, stroke=0)
        c.setStrokeColor(BORDER)
        c.setLineWidth(0.4)
        c.rect(x, y, w, h, fill=0, stroke=1)

    def txt(x, y, text, font="Helvetica", size=6, color=black, align="l"):
        """Escribe texto. align: l=left, c=center, r=right."""
        c.setFont(font, size)
        c.setFillColor(color)
        if align == "c":
            c.drawCentredString(x, y, str(text))
        elif align == "r":
            c.drawRightString(x, y, str(text))
        else:
            c.drawString(x, y, str(text))

    def txt_rot(x, y, text, font="Helvetica-Bold", size=5.5):
        """Escribe texto rotado 90° (de abajo hacia arriba)."""
        c.saveState()
        c.translate(x, y)
        c.rotate(90)
        c.setFont(font, size)
        c.setFillColor(black)
        c.drawString(1.5, -1, str(text))
        c.restoreState()

    # ══════════════════════════════════════════════════════════════════════
    # 1. ENCABEZADO INSTITUCIONAL
    # ══════════════════════════════════════════════════════════════════════

    y = H - 8 * mm
    th = 13 * mm  # alto del título
    y -= th

    s1w = CW * 0.28
    s2w = CW * 0.28
    s3w = CW * 0.44

    # Bloque 1: Gobierno Regional
    rect(ML, y, s1w, th, white)
    txt(ML + s1w / 2, y + th / 2 + 2, "GOBIERNO REGIONAL", "Helvetica-Bold", 9, black, "c")
    txt(ML + s1w / 2, y + th / 2 - 8, "DE AREQUIPA", "Helvetica-Bold", 9, black, "c")

    # Bloque 2: Cadena de Custodia
    x2 = ML + s1w
    rect(x2, y, s2w, th, white)
    txt(x2 + s2w / 2, y + th / 2 - 2, "CADENA DE CUSTODIA", "Helvetica-Bold", 11, black, "c")

    # Bloque 3: AUTODEMA / LVCA
    x3 = x2 + s2w
    rect(x3, y, s3w, th, white)
    txt(x3 + s3w / 2, y + th / 2 + 2, "AUTORIDAD AUTONOMA DE MAJES", "Helvetica-Bold", 8, black, "c")
    txt(x3 + s3w / 2, y + th / 2 - 7,
        "LABORATORIO DE VIGILANCIA Y CALIDAD DE AGUA (LVCA)", "Helvetica", 6.5, black, "c")

    # ══════════════════════════════════════════════════════════════════════
    # 2. DATOS GENERALES + PRESERVACIÓN
    # ══════════════════════════════════════════════════════════════════════

    bh = 5 * mm  # alto de barra
    y -= bh

    # Barra "DATOS GENERALES"
    dg_w = CW * 0.52
    rect(ML, y, dg_w, bh, GREEN)
    txt(ML + 3, y + 1.5 * mm, "DATOS GENERALES DEL MONITOREO", "Helvetica-Bold", 7)

    # Barra "PRESERVACIÓN"
    pr_x = ML + dg_w
    pr_w = CW * 0.28
    rect(pr_x, y, pr_w, bh, GREEN)
    txt(pr_x + 3, y + 1.5 * mm, "PRESERVACIÓN", "Helvetica-Bold", 7)

    # Barra "PÁGINA"
    pg_x = pr_x + pr_w
    pg_w = CW * 0.20
    rect(pg_x, y, pg_w, bh, white)
    txt(pg_x + 3, y + 1.5 * mm, "PAGINA: 1 de 1", "Helvetica", 6)

    # ── Filas de datos del encabezado ─────────────────────────────────────
    rh = 4.5 * mm

    # Marcas de preservación: derivadas de los parámetros seleccionados,
    # no de configuración manual. Si algún param usa HNO3, aparece con X.
    def _pres_mark(nombre: str) -> str:
        return "X" if nombre in preservantes_usados else "—"

    # Fila: Área + HNO3/H2SO4 + Código doc
    y -= rh
    rect(ML, y, dg_w, rh, white)
    txt(ML + 3, y + 1.2 * mm, f"ÁREA: {cfg.get('area', '')[:75]}", "Helvetica", 5.5)

    rect(pr_x, y, pr_w, rh, HexColor("#D5E8C0"))
    txt(pr_x + 3, y + 1.2 * mm,
        f"HNO3: {_pres_mark('HNO3')}     "
        f"H2SO4: {_pres_mark('H2SO4')}", "Helvetica", 5.5)

    rev_h = rh * 3
    rect(pg_x, y - rh * 2, pg_w, rev_h, white)
    txt(pg_x + 3, y + 1 * mm,
        f"Código: {cfg.get('codigo_documento', 'CC-MON-01')}  Rev: {cfg.get('revision', '03')}",
        "Helvetica", 5.5)
    txt(pg_x + 3, y + 1 * mm - rh, f"Fecha: {datetime.utcnow().strftime('%d-%b-%Y')}", "Helvetica", 5.5)

    # Fila: Dirección + HCl/Lugol/Formol
    y -= rh
    rect(ML, y, dg_w, rh, white)
    txt(ML + 3, y + 1.2 * mm,
        f"DIRECCIÓN: {cfg.get('direccion', '')}     TEL: {cfg.get('telefono', '')}",
        "Helvetica", 5.5)

    rect(pr_x, y, pr_w, rh, HexColor("#D5E8C0"))
    txt(pr_x + 3, y + 1.2 * mm,
        f"HCl: {_pres_mark('HCl')}     "
        f"Lugol: {_pres_mark('Lugol')}     "
        f"Formol: {_pres_mark('Formol')}     "
        f"S/P: {_pres_mark('S/P')}", "Helvetica", 5.5)

    # Fila: Institución + Urgencia
    y -= rh
    rect(ML, y, dg_w, rh, white)
    txt(ML + 3, y + 1.2 * mm,
        f"INSTITUCIÓN: {cfg.get('institucion', '')}     "
        f"URGENCIA: {cfg.get('urgencia', 'Regular')}", "Helvetica", 5.5)
    rect(pr_x, y, pr_w, rh, white)

    # Fila: Campaña + Cuenca
    campana_label = cfg.get("campana_label") or datos.get("lugar", "")
    y -= rh
    rect(ML, y, dg_w, rh, white)
    txt(ML + 3, y + 1.2 * mm,
        f"CAMPAÑA: {campana_label}     CUENCA: {datos.get('cuenca', '')}",
        "Helvetica-Bold", 5.5)
    rect(pr_x, y, pr_w + pg_w, rh, white)

    # ══════════════════════════════════════════════════════════════════════
    # 3. TABLA DE DATOS
    # ══════════════════════════════════════════════════════════════════════

    y -= 2 * mm

    # Definición de columnas (nombre, ancho) — misma estructura que Excel
    params_campo_pdf = list(_get_parametros_campo()[:_TEMPLATE_FIELD_COUNT])
    n_lab = len(params_lab)
    n_field = len(params_campo_pdf)

    fixed_cols = [
        ("N°", 4.5 * mm),
        ("Código Lab.", 18 * mm),
        ("Punto de Muestreo / Código", 40 * mm),
        ("Fecha", 14 * mm),
        ("Hora", 9 * mm),
        ("Matriz", 7 * mm),
        ("UTM Este", 14 * mm),
        ("UTM Norte", 14 * mm),
        ("Alt.", 9 * mm),
        ("Zona", 8 * mm),
        ("V", 4 * mm),
        ("P", 4 * mm),
    ]
    fixed_w = sum(w for _, w in fixed_cols)

    lab_col_w = 3.8 * mm
    field_col_w = 8.5 * mm
    lab_total = lab_col_w * n_lab
    field_total = field_col_w * n_field

    obs_w = CW - fixed_w - lab_total - field_total
    obs_w = max(obs_w, 15 * mm)

    # ── Fila de secciones ─────────────────────────────────────────────────
    sec_h = 5 * mm
    y -= sec_h
    x = ML

    rect(x, y, fixed_w, sec_h, GREEN)
    txt(x + fixed_w / 2, y + 1.5 * mm, "DATOS DE MUESTRA", "Helvetica-Bold", 6, black, "c")
    x += fixed_w

    rect(x, y, lab_total, sec_h, GREEN)
    txt(x + lab_total / 2, y + 1.5 * mm, "PARÁMETROS DE LABORATORIO", "Helvetica-Bold", 6, black, "c")
    x += lab_total

    rect(x, y, field_total, sec_h, LIGHT_GREEN)
    txt(x + field_total / 2, y + 1.5 * mm, "PARÁMETROS DE CAMPO", "Helvetica-Bold", 6, black, "c")
    x += field_total

    rect(x, y, obs_w, sec_h, GREEN)
    txt(x + obs_w / 2, y + 1.5 * mm, "OBS.", "Helvetica-Bold", 6, black, "c")

    # ── Fila de nombres de columna (rotados para params) ──────────────────
    hdr_h = 20 * mm
    y -= hdr_h
    x = ML

    # Columnas fijas — texto horizontal
    for name, w in fixed_cols:
        rect(x, y, w, hdr_h, GREEN)
        words = name.split()
        n_lines = len(words)
        for li, word in enumerate(words):
            ly = y + hdr_h / 2 + (n_lines / 2 - li - 0.5) * 6.5
            txt(x + w / 2, ly, word, "Helvetica-Bold", 5, black, "c")
        x += w

    # Params lab — texto rotado 90°, fondo blanco
    for p in params_lab:
        rect(x, y, lab_col_w, hdr_h, white)
        txt_rot(x + lab_col_w / 2 + 1.5, y + 1, p["nombre"])
        x += lab_col_w

    # Params campo — texto rotado 90°, fondo verde claro
    for p in params_campo_pdf:
        rect(x, y, field_col_w, hdr_h, LIGHT_GREEN)
        txt_rot(x + field_col_w / 2 + 1.5, y + 1, p["nombre"])
        x += field_col_w

    # Observaciones
    rect(x, y, obs_w, hdr_h, GREEN)
    txt(x + obs_w / 2, y + hdr_h / 2, "Observaciones", "Helvetica-Bold", 5, black, "c")

    # ── Filas de datos ────────────────────────────────────────────────────
    drh = 6 * mm  # alto de fila de datos

    for idx, m in enumerate(muestras):
        y -= drh
        if y < 45 * mm:
            break  # no hay espacio para más filas + pie

        pt = m.get("puntos_muestreo") or {}
        insitu = m.get("insitu", {})
        bg = white if idx % 2 == 0 else GRAY_BG

        tipo_mapa = {
            "laguna": "ADL", "rio": "ADR", "canal": "ADR",
            "manantial": "AMA", "embalse": "ADL", "pozo": "ASUB",
        }
        matriz = tipo_mapa.get(pt.get("tipo", ""), "AN")

        fecha_raw = m.get("fecha_muestreo", "")
        try:
            fecha_dt = datetime.strptime(str(fecha_raw)[:10], "%Y-%m-%d")
            fecha_str = fecha_dt.strftime("%d/%m/%y")
        except (ValueError, TypeError):
            fecha_str = str(fecha_raw)[:10]

        # Valores de las columnas fijas
        vals_fixed = [
            str(idx + 1),
            m.get("codigo", "") or "",
            f"{pt.get('nombre', '')[:28]} /{pt.get('codigo', '')}",
            fecha_str,
            (m.get("hora_recoleccion", "") or "")[:5],
            matriz,
            str(pt.get("utm_este", "") or ""),
            str(pt.get("utm_norte", "") or ""),
            str(pt.get("altitud_msnm", "") or ""),
            pt.get("utm_zona", "19L") or "",
            str(n_vidrio),
            str(n_plastico),
        ]

        x = ML
        for ci, (_, w) in enumerate(fixed_cols):
            rect(x, y, w, drh, bg)
            val = vals_fixed[ci]
            al = "l" if ci == 2 else "c"
            tx = x + 1.5 if al == "l" else x + w / 2
            txt(tx, y + 1.8 * mm, val[:int(w / 2.8)] if len(val) > int(w / 2.8) else val,
                "Helvetica", 5, black, al)
            x += w

        # Params lab — "x" en todas las columnas (solo aparecen los seleccionados)
        for _p in params_lab:
            rect(x, y, lab_col_w, drh, bg)
            txt(x + lab_col_w / 2, y + 1.8 * mm, "x", "Helvetica", 5, black, "c")
            x += lab_col_w

        # Params campo — valores
        for p in params_campo_pdf:
            rect(x, y, field_col_w, drh, bg)
            val = insitu.get(p["clave"])
            if val is not None:
                txt(x + field_col_w / 2, y + 1.8 * mm,
                    str(val).replace(".", ","), "Helvetica", 5, black, "c")
            x += field_col_w

        # Observaciones (auto-incluir clima, nivel, temp. transporte)
        obs_pdf_parts = []
        if m.get("clima"):
            obs_pdf_parts.append(f"Clima: {m['clima']}")
        if m.get("nivel_agua"):
            obs_pdf_parts.append(f"Nivel: {m['nivel_agua']}")
        if m.get("temperatura_transporte") is not None:
            obs_pdf_parts.append(f"T: {m['temperatura_transporte']}°C")
        if m.get("observaciones_campo"):
            obs_pdf_parts.append(m["observaciones_campo"])
        obs_text = "; ".join(obs_pdf_parts)[:40]
        rect(x, y, obs_w, drh, bg)
        txt(x + 1.5, y + 1.8 * mm, obs_text, "Helvetica", 4.5, black)

    # Filas vacías para completar mínimo visual
    remaining = max(0, 11 - len(muestras))
    for i in range(remaining):
        y -= drh
        if y < 45 * mm:
            break
        bg = white if (len(muestras) + i) % 2 == 0 else GRAY_BG
        x = ML
        for _, w in fixed_cols:
            rect(x, y, w, drh, bg)
            x += w
        for _ in params_lab:
            rect(x, y, lab_col_w, drh, bg)
            x += lab_col_w
        for _ in params_campo_pdf:
            rect(x, y, field_col_w, drh, bg)
            x += field_col_w
        rect(x, y, obs_w, drh, bg)

    # ══════════════════════════════════════════════════════════════════════
    # 4. PIE DE PÁGINA
    # ══════════════════════════════════════════════════════════════════════

    y -= 3 * mm
    pie_h = 6 * mm  # alto de fila de pie

    # Encabezados de sección
    sec_muestreo_w = CW * 0.30
    sec_recepcion_w = CW * 0.30
    sec_equipos_w = CW * 0.40

    y -= pie_h
    rect(ML, y, sec_muestreo_w, pie_h, GREEN)
    txt(ML + 3, y + 1.8 * mm, "MUESTREO REALIZADO POR:", "Helvetica-Bold", 6)

    rx = ML + sec_muestreo_w
    rect(rx, y, sec_recepcion_w, pie_h, GREEN)
    txt(rx + 3, y + 1.8 * mm, "RECEPCIÓN MUESTRA:", "Helvetica-Bold", 6)

    ex = rx + sec_recepcion_w
    rect(ex, y, sec_equipos_w, pie_h, GREEN)
    txt(ex + 3, y + 1.8 * mm, "DESCRIPCIÓN DE EQUIPOS UTILIZADOS:", "Helvetica-Bold", 6)

    # Detalle muestreo
    det_h = 5 * mm
    muestreo = cfg.get("muestreo_por", "laboratorio")

    y -= det_h
    rect(ML, y, sec_muestreo_w, det_h, white)
    txt(ML + 3, y + 1.5 * mm,
        f"Personal laboratorio: ({'X' if muestreo == 'laboratorio' else ' '})   "
        f"Otro: ({'X' if muestreo != 'laboratorio' else ' '})", "Helvetica", 5)

    rect(rx, y, sec_recepcion_w, det_h, white)
    txt(rx + 3, y + 1.5 * mm, f"Nombre: {cfg.get('nombre_receptor', '')}", "Helvetica", 5)

    equipos = cfg.get("equipos", EQUIPOS_DEFAULT)
    rect(ex, y, sec_equipos_w, det_h, white)
    if len(equipos) > 0:
        txt(ex + 3, y + 1.5 * mm,
            f"1. {equipos[0].get('codigo', '')} — {equipos[0].get('nombre', '')}", "Helvetica", 5)

    y -= det_h
    rect(ML, y, sec_muestreo_w, det_h, white)
    txt(ML + 3, y + 1.5 * mm, f"Nombre: {cfg.get('nombre_muestreador', '')}", "Helvetica", 5)

    rect(rx, y, sec_recepcion_w, det_h, white)
    txt(rx + 3, y + 1.5 * mm, "Fecha:", "Helvetica", 5)

    rect(ex, y, sec_equipos_w, det_h, white)
    if len(equipos) > 1:
        txt(ex + 3, y + 1.5 * mm,
            f"2. {equipos[1].get('codigo', '')} — {equipos[1].get('nombre', '')}", "Helvetica", 5)

    y -= det_h
    rect(ML, y, sec_muestreo_w, det_h, white)
    txt(ML + 3, y + 1.5 * mm, "Firma:", "Helvetica", 5)

    rect(rx, y, sec_recepcion_w, det_h, white)
    txt(rx + 3, y + 1.5 * mm, "Firma:", "Helvetica", 5)

    rect(ex, y, sec_equipos_w, det_h, white)

    # Supervisor
    y -= pie_h
    rect(ML, y, sec_muestreo_w, pie_h, GREEN)
    txt(ML + 3, y + 1.8 * mm,
        "COORDINADOR / SUPERVISOR:", "Helvetica-Bold", 6)

    rect(rx, y, sec_recepcion_w + sec_equipos_w, pie_h, white)
    txt(rx + 3, y + 1.8 * mm,
        f"Observaciones: {cfg.get('observaciones_generales', '')}", "Helvetica", 5)

    y -= det_h
    rect(ML, y, sec_muestreo_w, det_h, white)
    txt(ML + 3, y + 1.5 * mm, f"Nombre: {cfg.get('nombre_supervisor', '')}", "Helvetica", 5)

    rect(rx, y, sec_recepcion_w + sec_equipos_w, det_h, white)
    txt(rx + 3, y + 1.5 * mm,
        f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC — LVCA / AUTODEMA",
        "Helvetica", 5, HexColor("#888888"))

    # ── Condiciones de la muestra ─────────────────────────────────────────
    y -= pie_h
    cond = cfg.get("condiciones", {})
    rect(ML, y, CW, pie_h, HexColor("#FFF9E6"))
    cond_items = [
        ("Refrigerado", cond.get("refrigerado")),
        ("Icepack", cond.get("icepack")),
        ("Temp. ambiente", cond.get("temp_ambiente")),
        ("Congelado", cond.get("congelado")),
        ("Caja conservadora", cond.get("caja_conservadora")),
        ("Hielo potable", cond.get("hielo_potable")),
    ]
    cond_str = "CONDICIONES DE LA MUESTRA:   " + "    ".join(
        f"{name}: ({'X' if val else ' '})" for name, val in cond_items
    )
    txt(ML + 3, y + 1.8 * mm, cond_str, "Helvetica-Bold", 5.5)

    # ══════════════════════════════════════════════════════════════════════
    c.save()
    return output.getvalue()
