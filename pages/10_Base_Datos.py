"""
pages/10_Base_Datos.py
Base de Datos consolidada de resultados de monitoreo.

Tabla tipo hoja de cálculo con todos los resultados pivotados por parámetro.
- Filtros por campaña, punto, fecha
- Celdas coloreadas en rojo si el valor excede su ECA respectivo
- Edición directa para administradores
"""

from __future__ import annotations

from datetime import date, timedelta

import pandas as pd
import streamlit as st

from components.auth_guard import require_rol
from components.nav_context import consumir_contexto, ir_a, preseleccionar, rol_alcanza
from services.base_datos_service import (
    LIMITE_MUESTRAS,
    actualizar_resultado,
    crear_resultado,
    get_datos_consolidados,
    get_limites_eca_todos,
    get_parametros_map,
)

# Filas por página en la vista de tabla. La paginación solo se activa cuando
# hay más registros que esto; con conjuntos pequeños la tabla se muestra
# completa como antes. Reduce drásticamente el tamaño del DOM/HTML renderizado
# y el scroll inicial de la página.
_FILAS_POR_PAGINA = 100
from services.parametro_registry import (
    get_columnas_parametros,
    get_codigos_parametros,
    get_cat_params,
    get_lcm_por_codigo,
)
from components.ui_styles import aplicar_estilos, page_header, top_nav
from services.resultado_service import get_campanas
from services.punto_service import get_puntos
from services.muestra_service import actualizar_muestra
from services.audit_service import registrar_cambios_multiples
from services.logging_config import get_logger

logger = get_logger(__name__)


def _es_admin() -> bool:
    """Verifica si el usuario actual tiene rol de administrador."""
    sesion = st.session_state.get("sesion")
    if not sesion:
        return False
    return getattr(sesion, "rol", None) == "administrador"


# Decimales por categoría para mostrar en la tabla (no persiste en BD).
_FORMATO_POR_CATEGORIA = {
    "Parámetros de Campo": "%.2f",
    "Parámetros Físico-Químicos (Inorgánicos / Orgánicos)": "%.3f",
    "Parámetros Hidrobiológicos": "%.1f",
}
_FORMATO_FALLBACK = "%.4g"


def _formato_por_codigo(cat_params: dict) -> dict:
    """Mapa {codigo: format_string} según la categoría del parámetro."""
    mapa: dict[str, str] = {}
    for cat_nombre, codigos in cat_params.items():
        fmt = _FORMATO_POR_CATEGORIA.get(cat_nombre, _FORMATO_FALLBACK)
        for cod in codigos:
            mapa[cod] = fmt
    return mapa


def _excede_eca(valor, eca_id: str | None, param_codigo: str, limites: dict) -> bool:
    """Retorna True si el valor excede el ECA del punto."""
    if valor is None or eca_id is None:
        return False
    lim = limites.get((eca_id, param_codigo))
    if not lim:
        return False
    vmax = lim.get("valor_maximo")
    vmin = lim.get("valor_minimo")
    if vmax is not None and valor > vmax:
        return True
    if vmin is not None and valor < vmin:
        return True
    return False


def _colorear_celda(val, eca_id, param_codigo, limites):
    """Retorna estilo CSS si excede ECA."""
    if val is None or pd.isna(val):
        return ""
    try:
        v = float(val)
    except (ValueError, TypeError):
        return ""
    if _excede_eca(v, eca_id, param_codigo, limites):
        return "background-color: #fee2e2; color: #b91c1c; font-weight: bold;"
    return ""


# Estilos CSS para la tabla HTML con separadores amarillos por campaña.
# Paleta alineada con el tema global (ui_styles): header gris #f8fafc,
# solo bordes horizontales, zebra sutil y celda de excedencia con los
# mismos tonos del chip ECA "excede".
_BD_TABLE_CSS = """
<style>
  .bd-table-wrap {
    overflow: auto;
    max-height: 72vh;
    position: relative;
    border: 1px solid var(--lvca-border);
    border-radius: var(--lvca-radius-md);
    margin-bottom: 0.75rem;
    box-shadow: var(--lvca-shadow-xs);
    scrollbar-width: thin;
  }
  table.bd-table {
    border-collapse: separate;
    border-spacing: 0;
    font-size: 12px;
    width: max-content;
    min-width: 100%;
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    font-variant-numeric: tabular-nums;
  }
  /* box-sizing explícito: los offsets `left` de las columnas congeladas asumen
     que el ancho fijado (min/max-width) incluye el padding. Streamlit no aplica
     un reset border-box global, así que se fuerza aquí para que el apilado de
     las columnas fijas cuadre exactamente. */
  table.bd-table th,
  table.bd-table td { box-sizing: border-box; }
  table.bd-table thead th {
    position: sticky;
    top: 0;
    background: var(--lvca-surface-alt);
    color: #475569;
    font-weight: 600;
    font-size: 0.72rem;
    text-transform: uppercase;
    letter-spacing: 0.04em;
    padding: 8px 10px;
    border: none;
    border-bottom: 1px solid var(--lvca-border);
    text-align: center;
    white-space: nowrap;
    z-index: 3;
  }
  table.bd-table tbody td {
    padding: 5px 10px;
    border: none;
    border-bottom: 1px solid var(--lvca-border-soft);
    background: var(--lvca-bg-card);
    white-space: nowrap;
    text-align: right;
    color: var(--lvca-text);
  }
  table.bd-table tbody td.text { text-align: left; }
  table.bd-table tbody tr:nth-child(even) td { background: #fcfdfe; }
  table.bd-table tbody tr:hover td { background: #eef2f7; }
  table.bd-table td.exceed {
    background: #fee2e2 !important;
    color: #b91c1c;
    font-weight: 700;
  }

  /* ── Columnas identificadoras congeladas (freeze panes tipo Excel) ──
     Se fijan las primeras columnas estables: Fecha, Hora, Cód. Punto y Punto,
     con anchura fija para que el desplazamiento horizontal no oculte la
     estación. Fondo sólido para que el contenido que pasa por debajo no se
     transparente. */
  table.bd-table th.freeze,
  table.bd-table td.freeze {
    position: sticky;
    z-index: 2;
    overflow: hidden;
    text-overflow: ellipsis;
  }
  table.bd-table thead th.freeze { z-index: 4; white-space: normal; line-height: 1.15; }
  table.bd-table tbody td.freeze { background: var(--lvca-bg-card); }
  table.bd-table tbody tr:nth-child(even) td.freeze { background: #fcfdfe; }
  table.bd-table tbody tr:hover td.freeze { background: #eef2f7; }
  /* Anchos (border-box, incluyen padding) y offsets left acumulados:
     0 · 112 · 112+60=172 · 172+106=278. La columna Fecha (f0) es lo bastante
     ancha para mostrar la fecha ISO completa (2026-03-15) sin recortarla. */
  .bd-table .f0 { left: 0;     min-width: 112px; max-width: 112px; }
  .bd-table .f1 { left: 112px; min-width: 60px;  max-width: 60px;  }
  .bd-table .f2 { left: 172px; min-width: 106px; max-width: 106px; }
  .bd-table .f3 {
    left: 278px; min-width: 176px; max-width: 176px;
    box-shadow: 6px 0 6px -4px rgba(15, 23, 42, 0.14);
  }

  table.bd-table tr.bd-sep td {
    background: #fef3c7 !important;
    color: #92400e;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    padding: 7px 10px;
    border-top: 2px solid var(--lvca-acento-amarillo);
    border-bottom: 1px solid #fde68a;
    text-align: left;
  }
  /* La etiqueta de campaña permanece visible al desplazarse a la derecha. */
  table.bd-table tr.bd-sep .bd-sep-label {
    position: sticky;
    left: 12px;
    display: inline-block;
  }
</style>
"""

# Columnas identificadoras que se congelan a la izquierda (orden = prefijo real
# de la tabla). Son las que SIEMPRE están presentes, por lo que el índice de
# congelado es estable aunque se oculten "Código Lab." o "Profundidad".
_FREEZE_COLS = ["Fecha", "Hora", "Código Punto", "Punto"]
_FREEZE_INDEX = {name: i for i, name in enumerate(_FREEZE_COLS)}

# ── Abreviaturas de encabezado (SOLO presentación en pantalla) ───────────────
# El DataFrame y las descargas conservan los nombres completos; estas
# abreviaturas se aplican únicamente al texto del <th>, con el nombre completo
# disponible en el tooltip (atributo title). Alineadas con los `nombre_corto`
# del seed de parámetros, con los ajustes pedidos (T°, C.E., formas químicas).
ABREVIATURAS_PARAM: dict[str, str] = {
    # Campo
    "P001": "pH",
    "P002": "T°",
    "P003": "C.E.",
    "P004": "OD",
    "P006": "Turbidez",
    # Fisicoquímicos
    "P010": "Color",
    "P019": "DBO5",
    "P025": "Dureza",
    "P028": "SST",
    "P031": "NO3",
    "P032": "NO2",
    "P033": "NH3 total",
    "P034": "NH3 libre",
    "P036": "P total",
    "P038": "PO4",
    "P041": "SO4",
    "P042": "Cl",
    "P074": "Fe",
    "P077": "Mn",
    "P091": "Microcistina",
    "P124": "Cl-a",
    # Hidrobiológicos
    "P120": "Fitopl.",
    "P126": "Zoopl.",
    "P130": "Perifiton",
    "FITO_CYANOBACTERIA_CEL": "Cyanobacteria",
    "FITO_CYANOBACTERIA_BIOVOL": "Cyanobacteria biovol.",
    "FITO_BACILLARIOPHYTA": "Bacillariophyta",
    "FITO_CHLOROPHYTA": "Chlorophyta",
    "FITO_OCHROPHYTA": "Ochrophyta",
    "FITO_CHAROPHYTA": "Charophyta",
    "FITO_EUGLENOPHYTA": "Euglenophyta",
    "FITO_DINOPHYTA": "Dinophyta",
    "FITO_CRYPTOPHYTA": "Cryptophyta",
}

# Columnas identificadoras que conviene acortar; el resto (Fecha, Hora, Punto,
# Cuenca, Tipo, ECA) ya son cortas y se muestran igual.
ABREVIATURAS_FIJAS: dict[str, str] = {
    "Código Punto": "Cód. Punto",
    "Código Muestra": "Cód. Muestra",
    "Código Lab.": "Cód. Lab.",
    "Profundidad (m)": "Prof. (m)",
}


def _unidad_corta(label_completo: str) -> str:
    """Extrae la unidad (último paréntesis del label) y la simplifica.

    Los nombres pueden llevar paréntesis propios (p. ej.
    'Nitrógeno amoniacal total (N-NH3) (mg N-NH4/L)'), por lo que la unidad es
    SIEMPRE el último grupo entre paréntesis. Cualquier unidad de masa por
    litro ('mg N-NO3/L', 'mg CaCO3/L', 'mg O2/L'…) se colapsa a 'mg/L'; las
    demás unidades se mantienen tal cual.
    """
    ini = label_completo.rfind("(")
    fin = label_completo.rfind(")")
    if ini == -1 or fin == -1 or fin < ini:
        return ""
    unidad = label_completo[ini + 1:fin].strip()
    if unidad.startswith("mg") and unidad.endswith("/L"):
        return "mg/L"
    return unidad


def _fmt_valor(val, fmt: str) -> str:
    """Formatea un valor numérico usando un format string estilo %.2f."""
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    try:
        return fmt % float(val)
    except (TypeError, ValueError):
        return str(val)


def _hora_a_minutos(h) -> int:
    """Convierte una hora 'HH:MM' a minutos desde medianoche, para ordenar
    cronológicamente. Las muestras sin hora (o con formato no reconocible) se
    ordenan al final del día, no al principio."""
    if not h:
        return 24 * 60 + 1  # sin hora → al final del día
    s = str(h).strip()
    try:
        partes = s.split(":")
        hh = int(partes[0])
        mm = int(partes[1]) if len(partes) > 1 else 0
        if 0 <= hh <= 23 and 0 <= mm <= 59:
            return hh * 60 + mm
    except (ValueError, IndexError):
        pass
    return 24 * 60 + 2  # formato raro → después incluso de las sin hora


_MES_ES = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
    7: "JULIO", 8: "AGOSTO", 9: "SETIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}


def _etiqueta_campana(d: dict) -> str:
    """Etiqueta amarilla del separador: MES YYYY · CODIGO — nombre."""
    fecha = d.get("campana_fecha_inicio") or d.get("fecha") or ""
    mes_txt = ""
    if fecha and len(fecha) >= 7:
        try:
            anio, mes = int(fecha[:4]), int(fecha[5:7])
            mes_txt = f"{_MES_ES.get(mes, '')} {anio}"
        except ValueError:
            mes_txt = ""
    cod = d.get("campana_codigo") or ""
    nom = d.get("campana_nombre") or ""
    partes = [p for p in (mes_txt, cod, nom) if p]
    return "  ·  ".join(partes) if partes else "Sin campaña"


def _render_tabla_por_campana(
    df: pd.DataFrame,
    datos: list[dict],
    columnas_visibles: list[tuple[str, str]],
    formato_codigo: dict,
    limites: dict,
    lcm_codigo: dict | None = None,
) -> str:
    """
    Renderiza la base de datos como tabla HTML con separadores amarillos
    por campaña (estilo Excel: una fila completa amarilla con MES + código).

    El orden de las filas debe coincidir entre `df` y `datos`.
    """
    from html import escape

    columnas = list(df.columns)
    label_to_codigo = {label: cod for cod, label in columnas_visibles}
    text_cols = {"Fecha", "Hora", "Código Punto", "Punto", "Código Muestra",
                 "Código Lab.", "Cuenca", "Tipo", "ECA"}

    # Cabecera (con clases de congelado en las columnas identificadoras).
    # El texto se abrevia SOLO para mostrar; el nombre completo original queda
    # en el tooltip (title). Los datos y las descargas usan el label completo.
    thead_cells = []
    for c in columnas:
        fi = _FREEZE_INDEX.get(c)
        cls = f' class="freeze f{fi}"' if fi is not None else ""
        cod = label_to_codigo.get(c)
        if cod is not None:
            abrev = ABREVIATURAS_PARAM.get(cod) or c.split(" (")[0]
            unidad = _unidad_corta(c)
            corto = f"{abrev} ({unidad})" if unidad and unidad != abrev else abrev
        else:
            corto = ABREVIATURAS_FIJAS.get(c, c)
        titulo = f' title="{escape(c)}"' if corto != c else ""
        thead_cells.append(f"<th{cls}{titulo}>{escape(corto)}</th>")
    thead = "".join(thead_cells)

    # Cuerpo: agrupar filas consecutivas que comparten campana_id.
    body_parts: list[str] = []
    ultimo_campana_key = object()
    n_cols = len(columnas)

    for idx, d in enumerate(datos):
        campana_key = d.get("campana_id") or d.get("campana_codigo") or "__sin__"
        if campana_key != ultimo_campana_key:
            label = escape(_etiqueta_campana(d))
            body_parts.append(
                f'<tr class="bd-sep"><td colspan="{n_cols}">'
                f'<span class="bd-sep-label">{label}</span></td></tr>'
            )
            ultimo_campana_key = campana_key

        eca_id = d.get("eca_id")
        celdas: list[str] = []
        fila = df.iloc[idx]
        for col in columnas:
            raw = fila[col]
            cod = label_to_codigo.get(col)
            fi = _FREEZE_INDEX.get(col)
            clases: list[str] = []
            if cod is not None:
                fmt = formato_codigo.get(cod, _FORMATO_FALLBACK)
                es_num = raw is not None and not (isinstance(raw, float) and pd.isna(raw))
                lcm = (lcm_codigo or {}).get(cod)
                bajo_lcm = False
                if es_num and lcm is not None:
                    try:
                        bajo_lcm = float(raw) < lcm
                    except (TypeError, ValueError):
                        bajo_lcm = False
                if bajo_lcm:
                    # Por debajo del Límite de Cuantificación: se reporta como
                    # '< LCM' (no cuantificable) y nunca excede el ECA.
                    txt = f"< {_fmt_valor(lcm, fmt)}"
                else:
                    txt = _fmt_valor(raw, fmt)
                    if es_num and _excede_eca(raw, eca_id, cod, limites):
                        clases.append("exceed")
            else:
                if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                    txt = ""
                elif col == "Profundidad (m)":
                    try:
                        txt = f"{float(raw):.2f}"
                    except (TypeError, ValueError):
                        txt = str(raw)
                else:
                    txt = str(raw)
                if col in text_cols:
                    clases.append("text")
            titulo = ""
            if fi is not None:
                clases.extend(("freeze", f"f{fi}"))
                if txt:
                    titulo = f' title="{escape(str(txt))}"'
            cls = f' class="{" ".join(clases)}"' if clases else ""
            celdas.append(f"<td{cls}{titulo}>{escape(txt)}</td>")
        body_parts.append("<tr>" + "".join(celdas) + "</tr>")

    return (
        _BD_TABLE_CSS
        + '<div class="bd-table-wrap"><table class="bd-table">'
        + f"<thead><tr>{thead}</tr></thead>"
        + "<tbody>"
        + "".join(body_parts)
        + "</tbody></table></div>"
    )


@st.cache_data(show_spinner=False)
def _df_a_csv(df: pd.DataFrame) -> bytes:
    """Serializa el DataFrame a CSV, cacheado por contenido. Evita re-serializar
    el conjunto completo en cada rerun (paginación, filtros de presentación), ya
    que st.download_button evalúa su argumento `data` siempre.

    Se codifica en UTF-8 **con BOM** (utf-8-sig) para que Excel abra el archivo
    con las tildes y la ñ correctas (sin BOM Excel asume Latin-1 y muestra
    caracteres corruptos como 'Ã³' o 'Ã±')."""
    return df.to_csv(index=False).encode("utf-8-sig")


# ── Exportación a Excel (.xlsx) con formato legible ──────────────────────────

def _num_format_excel(fmt: str) -> str:
    """Traduce un format string de Python (%.2f) al formato numérico de Excel."""
    mapa = {"%.1f": "0.0", "%.2f": "0.00", "%.3f": "0.000", "%.4f": "0.0000"}
    if fmt in mapa:
        return mapa[fmt]
    if "g" in (fmt or "").lower():
        return "0.####"
    return "0.00"


def _construir_xlsx(
    df: pd.DataFrame,
    datos: list[dict],
    columnas_visibles: list[tuple[str, str]],
    formato_codigo: dict,
    limites: dict,
    lcm_codigo: dict | None = None,
) -> bytes:
    """Genera un .xlsx con formato legible que replica la vista en pantalla:
    encabezado destacado, anchos de columna adecuados, separadores amarillos por
    campaña, celdas de excedencia en rojo, '< LCM' para no cuantificables y panel
    congelado (encabezado + columnas identificadoras). Al ser .xlsx nativo, las
    tildes y la ñ se ven siempre correctas."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    columnas = list(df.columns)
    label_to_codigo = {label: cod for cod, label in columnas_visibles}
    text_cols = {"Fecha", "Hora", "Código Punto", "Punto", "Código Muestra",
                 "Código Lab.", "Cuenca", "Tipo", "ECA"}
    n_cols = len(columnas)

    wb = Workbook()
    ws = wb.active
    ws.title = "Base de Datos"

    header_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sep_font = Font(name="Calibri", bold=True, size=10, color="7C4A03")
    sep_fill = PatternFill("solid", fgColor="FDE68A")
    exceed_fill = PatternFill("solid", fgColor="FEE2E2")
    exceed_font = Font(name="Calibri", bold=True, size=10, color="B91C1C")
    normal_font = Font(name="Calibri", size=10)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="E2E8F0")
    borde = Border(bottom=thin)

    # Encabezado
    for j, col in enumerate(columnas, start=1):
        c = ws.cell(row=1, column=j, value=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = borde
    ws.row_dimensions[1].height = 30

    row_idx = 2
    ultimo_campana_key = object()
    for i, d in enumerate(datos):
        campana_key = d.get("campana_id") or d.get("campana_codigo") or "__sin__"
        if campana_key != ultimo_campana_key:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=n_cols)
            sc = ws.cell(row=row_idx, column=1, value=_etiqueta_campana(d))
            sc.font = sep_font
            sc.alignment = left_align
            for j in range(1, n_cols + 1):
                ws.cell(row=row_idx, column=j).fill = sep_fill
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
            ultimo_campana_key = campana_key

        eca_id = d.get("eca_id")
        fila = df.iloc[i]
        for j, col in enumerate(columnas, start=1):
            raw = fila[col]
            cod = label_to_codigo.get(col)
            cell = ws.cell(row=row_idx, column=j)
            cell.font = normal_font
            cell.border = borde
            if cod is not None:
                es_num = raw is not None and not (isinstance(raw, float) and pd.isna(raw))
                lcm = (lcm_codigo or {}).get(cod)
                bajo_lcm = False
                if es_num and lcm is not None:
                    try:
                        bajo_lcm = float(raw) < lcm
                    except (TypeError, ValueError):
                        bajo_lcm = False
                cell.alignment = right_align
                if bajo_lcm:
                    cell.value = f"< {_fmt_valor(lcm, formato_codigo.get(cod, _FORMATO_FALLBACK))}"
                elif es_num:
                    cell.value = float(raw)
                    cell.number_format = _num_format_excel(
                        formato_codigo.get(cod, _FORMATO_FALLBACK)
                    )
                    if _excede_eca(raw, eca_id, cod, limites):
                        cell.fill = exceed_fill
                        cell.font = exceed_font
            else:
                if raw is None or (isinstance(raw, float) and pd.isna(raw)):
                    cell.value = None
                elif col == "Profundidad (m)":
                    try:
                        cell.value = float(raw)
                        cell.number_format = "0.00"
                    except (TypeError, ValueError):
                        cell.value = str(raw)
                    cell.alignment = right_align
                else:
                    cell.value = str(raw)
                    cell.alignment = left_align if col in text_cols else right_align
        row_idx += 1

    # Anchos de columna: base por tipo + ajuste al contenido real (con tope),
    # para que los valores y nombres de estación se lean sin quedar cortados.
    for j, col in enumerate(columnas, start=1):
        cod = label_to_codigo.get(col)
        if cod is not None:
            base = 12
        elif col == "Punto":
            base = 26
        elif col in ("Código Muestra", "Código Lab.", "Cuenca"):
            base = 16
        else:
            base = 11
        try:
            max_data = max((len(str(v)) for v in df[col] if v is not None), default=0)
        except Exception:
            max_data = 0
        w = max(base, len(str(col)) + 2, min(38, max_data + 2))
        ws.column_dimensions[get_column_letter(j)].width = w

    # Congelar encabezado + columnas identificadoras (hasta 'Punto' inclusive).
    # No se aplica auto_filter: la hoja intercala filas de separador fusionadas
    # por campaña y un autofiltro sobre ellas se comporta de forma inconsistente.
    freeze_col = 1
    for k, col in enumerate(columnas, start=1):
        if col == "Punto":
            freeze_col = k + 1
            break
    # Coordenada como texto (no la celda): la fila 2 suele ser un separador de
    # campaña fusionado y ws.cell() devolvería un MergedCell no admitido aquí.
    ws.freeze_panes = f"{get_column_letter(freeze_col)}2"

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


@st.cache_data(show_spinner=False)
def _xlsx_cacheado(sig: str, _df, _datos, _columnas_visibles, _formato_codigo, _limites, _lcm):
    """Envuelve _construir_xlsx con caché por `sig` (firma de filtros + conteos).
    Los argumentos con guion bajo no se hashean; se limpia explícitamente tras
    cada edición para no servir un Excel obsoleto."""
    return _construir_xlsx(_df, _datos, _columnas_visibles, _formato_codigo, _limites, _lcm)


# ── Edición en cuadrícula (compartida por la vista inline y la pestaña) ───────

# Metadatos editables de una muestra: (etiqueta visible, clave en `datos`,
# columna real en la tabla `muestras`, tipo).
_META_EDITABLES = [
    ("Hora",       "hora",                "hora_recoleccion",  "text"),
    ("Prof. (m)",  "profundidad",         "profundidad_valor", "num"),
    ("Cód. Lab.",  "codigo_laboratorio",  "codigo_laboratorio", "text"),
    ("Clima",      "clima",               "clima",             "text"),
    ("Nivel agua", "nivel_agua",          "nivel_agua",        "text"),
]

_ID_COLS_EDITOR = ["Fecha", "Código Punto", "Punto", "Código Muestra", "ECA"]


def _es_vacio(x) -> bool:
    return x is None or x == "" or (isinstance(x, float) and pd.isna(x))


def _celda_distinta_num(old, new) -> bool:
    """True si el valor numérico de la celda cambió (tolerante a None/NaN)."""
    o_e, n_e = _es_vacio(old), _es_vacio(new)
    if o_e and n_e:
        return False
    if o_e != n_e:
        return True
    try:
        return abs(float(old) - float(new)) > 1e-9
    except (TypeError, ValueError):
        return str(old) != str(new)


def _sanitizar_formato_number(fmt: str) -> str:
    """st.column_config.NumberColumn no soporta de forma fiable %g → usar %.4f."""
    if not fmt or "g" in fmt.lower():
        return "%.4f"
    return fmt


def _usuario_id_actual() -> str | None:
    """Identificador del usuario para la auditoría (id interno si es resoluble,
    si no el uid de Auth). audit_log.usuario_id es TEXT, así que cualquiera vale."""
    sesion = st.session_state.get("sesion")
    uid = getattr(sesion, "uid", None) if sesion else None
    if not uid:
        return None
    try:
        from services.resultado_service import _get_usuario_interno_id
        interno = _get_usuario_interno_id(uid)
        if interno:
            return interno
    except Exception:
        pass
    return uid


def _limpiar_caches_pagina() -> None:
    """Limpia las cachés st.cache_data locales de la página que dependen de los
    datos editados, para que la vista y las descargas reflejen los cambios
    inmediatamente (el backend ya invalidó las cachés @cached del servicio)."""
    for fn in (_df_a_csv, _xlsx_cacheado):
        try:
            fn.clear()
        except Exception:
            logger.warning("No se pudo limpiar la caché de página %s tras editar; "
                           "podrían quedar datos obsoletos hasta el próximo TTL.",
                           getattr(fn, "__name__", fn), exc_info=True)


def _render_grid_editor(
    datos_subset: list[dict],
    columnas_visibles: list[tuple[str, str]],
    formato_codigo: dict,
    limites: dict,
    param_map: dict,
    key_prefix: str,
) -> None:
    """Cuadrícula editable (tipo Excel) para un conjunto de muestras.

    Columnas identificadoras bloqueadas; metadatos y valores de parámetros
    editables. Al guardar, cada cambio se persiste vía el servicio (misma tabla
    que Resultados e Informes) y se audita. Reutilizada por la edición inline de
    la vista y por la pestaña "Edición de datos" (filtrada por campaña)."""
    from html import escape

    if not datos_subset:
        st.info("No hay muestras para editar en esta selección.")
        return

    # ── Configuración de columnas (común a todos los bloques) ────────────
    cfg = {
        "Fecha": st.column_config.TextColumn("Fecha", disabled=True, width="small"),
        "Código Punto": st.column_config.TextColumn("Cód. Punto", disabled=True, width="small"),
        "Punto": st.column_config.TextColumn("Punto", disabled=True, width="medium"),
        "Código Muestra": st.column_config.TextColumn("Cód. Muestra", disabled=True, width="small"),
        "ECA": st.column_config.TextColumn("ECA", disabled=True, width="small"),
        "Hora": st.column_config.TextColumn("Hora", help="HH:MM"),
        "Prof. (m)": st.column_config.NumberColumn("Prof. (m)", format="%.2f",
                                                    help="Profundidad de muestreo (m)"),
        "Cód. Lab.": st.column_config.TextColumn("Cód. Lab."),
        "Clima": st.column_config.TextColumn("Clima"),
        "Nivel agua": st.column_config.TextColumn("Nivel agua"),
    }
    for cod, label in columnas_visibles:
        cfg[label] = st.column_config.NumberColumn(
            label, format=_sanitizar_formato_number(formato_codigo.get(cod, _FORMATO_FALLBACK))
        )

    def _build_df(sub: list[dict]) -> pd.DataFrame:
        filas = []
        for d in sub:
            fila = {
                "Fecha": d.get("fecha", ""),
                "Código Punto": d.get("punto_codigo", ""),
                "Punto": d.get("punto_nombre", ""),
                "Código Muestra": d.get("codigo_muestra", ""),
                "ECA": d.get("eca_codigo", ""),
                "Hora": d.get("hora", "") or "",
                "Prof. (m)": d.get("profundidad"),
                "Cód. Lab.": d.get("codigo_laboratorio", "") or "",
                "Clima": d.get("clima", "") or "",
                "Nivel agua": d.get("nivel_agua", "") or "",
            }
            for cod, label in columnas_visibles:
                fila[label] = d.get(cod)
            filas.append(fila)
        df_e = pd.DataFrame(filas)
        # Forzar dtype numérico en columnas numéricas (evita quejas de NumberColumn).
        for _c in ["Prof. (m)"] + [label for _cod, label in columnas_visibles]:
            if _c in df_e.columns:
                df_e[_c] = pd.to_numeric(df_e[_c], errors="coerce")
        return df_e

    # ── Agrupar por campaña (los datos ya vienen ordenados por campaña) ──
    # Cada bloque lleva una banda amarilla con MES · CÓDIGO, igual que los
    # separadores de la vista de consulta, para diferenciar mejor al editar.
    grupos: list[tuple[str, list[int]]] = []
    ultimo = object()
    for i, d in enumerate(datos_subset):
        k = d.get("campana_id") or d.get("campana_codigo") or "__sin__"
        if k != ultimo:
            grupos.append((_etiqueta_campana(d), []))
            ultimo = k
        grupos[-1][1].append(i)

    _sep_style = ("background:#fef3c7;color:#92400e;font-weight:700;"
                  "text-transform:uppercase;letter-spacing:.5px;padding:7px 12px;"
                  "border-top:2px solid #F59E0B;border-bottom:1px solid #fde68a;"
                  "border-radius:6px 6px 0 0;font-size:.78rem;margin:.6rem 0 -.3rem;")

    editores: list[tuple[list[int], pd.DataFrame]] = []
    for gi, (label, idxs) in enumerate(grupos):
        st.markdown(f"<div style='{_sep_style}'>{escape(label)}</div>",
                    unsafe_allow_html=True)
        sub = [datos_subset[i] for i in idxs]
        edited = st.data_editor(
            _build_df(sub),
            column_config=cfg,
            hide_index=True,
            num_rows="fixed",
            use_container_width=True,
            height=min(520, 80 + 35 * len(sub)),
            key=f"{key_prefix}_g{gi}_editor",
        )
        editores.append((idxs, edited))

    c1, c2 = st.columns([1.2, 3.8])
    with c1:
        guardar = st.button(
            "Guardar cambios", type="primary", key=f"{key_prefix}_guardar",
            icon=":material/save:", use_container_width=True,
        )
    with c2:
        st.caption("Edita las celdas y pulsa **Guardar cambios**. Los valores y "
                   "metadatos se actualizan en **Resultados** e **Informes** al instante.")

    if not guardar:
        return

    uid = _usuario_id_actual()
    n_val = 0
    n_meta = 0
    errores: list[str] = []

    for idxs, edited in editores:
        for r, gi in enumerate(idxs):
            d = datos_subset[gi]
            cod_muestra = d.get("codigo_muestra", "?")

            # 1) Metadatos de la muestra
            meta_upd: dict = {}
            cambios_aud: dict = {}
            for lbl, dkey, bdfield, tipo in _META_EDITABLES:
                if lbl not in edited.columns:
                    continue
                old = d.get(dkey)
                new = edited.at[r, lbl]
                if tipo == "num":
                    if _celda_distinta_num(old, new):
                        val = None if _es_vacio(new) else float(new)
                        meta_upd[bdfield] = val
                        cambios_aud[bdfield] = (str(old) if old is not None else None,
                                                str(val) if val is not None else None)
                else:
                    old_s = "" if _es_vacio(old) else str(old).strip()
                    new_s = "" if _es_vacio(new) else str(new).strip()
                    if old_s != new_s:
                        meta_upd[bdfield] = new_s or None
                        cambios_aud[bdfield] = (old_s or None, new_s or None)
            if meta_upd:
                try:
                    actualizar_muestra(d["muestra_id"], meta_upd)
                    registrar_cambios_multiples("muestras", d["muestra_id"], "editar",
                                                cambios_aud, usuario_id=uid)
                    n_meta += 1
                except Exception as e:
                    errores.append(f"{cod_muestra} (metadatos): {e}")

            # 2) Valores de parámetros
            resultado_ids = d.get("_resultado_ids", {})
            for cod, label in columnas_visibles:
                if label not in edited.columns:
                    continue
                old = d.get(cod)
                new = edited.at[r, label]
                if not _celda_distinta_num(old, new):
                    continue
                new_val = None
                if not _es_vacio(new):
                    try:
                        new_val = float(new)
                    except (TypeError, ValueError):
                        new_val = None
                try:
                    info = resultado_ids.get(cod)
                    if info:
                        actualizar_resultado(info["resultado_id"], new_val, usuario_id=uid)
                    elif new_val is not None:
                        pid = param_map.get(cod)
                        if not pid:
                            errores.append(f"{cod_muestra}/{cod}: parámetro sin id")
                            continue
                        crear_resultado(d["muestra_id"], pid, new_val, usuario_id=uid)
                    else:
                        continue  # se vació una celda que ya estaba vacía
                    n_val += 1
                except Exception as e:
                    errores.append(f"{cod_muestra}/{cod}: {e}")

    for err in errores:
        st.error(f"Error al guardar {err}")

    if n_val or n_meta:
        _limpiar_caches_pagina()
        # Contador de versión: garantiza que la firma del Excel/CSV cacheado
        # cambie tras cada guardado, aunque el conteo de valores no varíe y
        # aunque un .clear() fallara silenciosamente.
        st.session_state["bd_data_version"] = st.session_state.get("bd_data_version", 0) + 1
        st.success(
            f"Guardado: {n_val} valor(es) y {n_meta} muestra(s) con metadatos "
            "actualizados. Los cambios ya se reflejan en Resultados e Informes."
        )
        st.rerun()
    elif not errores:
        st.info("No se detectaron cambios.")


@require_rol("visitante")
def main() -> None:
    aplicar_estilos()
    top_nav()
    page_header(
        "Base de Datos",
        "Consolidado de datos de campo, fisicoquímicos e hidrobiológicos por campaña",
    )

    es_admin = _es_admin()

    # Contexto de navegación: otra página pidió filtrar por campaña/punto
    ctx = consumir_contexto("base_datos")

    # ── Filtros (una sola fila, en main area) ───────────────────────────
    fc1, fc2, fc3, fc4, fc5 = st.columns([1.3, 1.5, 1.2, 0.9, 0.9])
    with fc1:
        campanas = get_campanas()
        opciones_camp = {"Todas las campañas": None}
        opciones_camp.update({
            f"{c['codigo']} — {c['nombre']}": c["id"] for c in campanas
        })
        preseleccionar("bd_camp", opciones_camp, ctx.get("campana_id"))
        sel_camp = st.selectbox("Campaña", list(opciones_camp.keys()), key="bd_camp")
        campana_id = opciones_camp[sel_camp]
    with fc2:
        # Filtro por nombre del lugar (represa / río / bocatoma).
        # Un mismo "lugar" puede tener varios puntos físicos: al seleccionarlo
        # se filtra por TODOS los punto_ids que comparten ese nombre.
        puntos = get_puntos(solo_activos=True)
        lugares: dict[str, list[str]] = {}
        for p in puntos:
            nombre = (p.get("nombre") or p.get("codigo") or "").strip()
            if not nombre:
                continue
            lugares.setdefault(nombre, []).append(p["id"])

        # Etiqueta con el tipo entre paréntesis cuando es único, para distinguir
        # "Río Sumbay" de "Represa Frayle" visualmente.
        tipos_por_nombre: dict[str, set[str]] = {}
        for p in puntos:
            nombre = (p.get("nombre") or p.get("codigo") or "").strip()
            if nombre:
                tipos_por_nombre.setdefault(nombre, set()).add((p.get("tipo") or "").strip())

        opciones_lugar: dict[str, tuple[str, ...] | None] = {"Todos los lugares": None}
        for nombre in sorted(lugares.keys()):
            tipos = {t for t in tipos_por_nombre.get(nombre, set()) if t}
            sufijo = f"  ·  {next(iter(tipos)).capitalize()}" if len(tipos) == 1 else ""
            opciones_lugar[f"{nombre}{sufijo}"] = tuple(lugares[nombre])

        # Pre-selección por punto_id del contexto: el "lugar" que lo contiene
        if ctx.get("punto_id"):
            for _lbl, _ids in opciones_lugar.items():
                if _ids and ctx["punto_id"] in _ids:
                    st.session_state["bd_lugar"] = _lbl
                    break

        sel_lugar = st.selectbox(
            "Lugar de muestreo",
            list(opciones_lugar.keys()),
            key="bd_lugar",
            help="Filtra por nombre del lugar (represa, río, bocatoma…). "
                 "Muestra todas las muestras de los puntos vinculados al lugar.",
        )
        punto_ids_filtro = opciones_lugar[sel_lugar]
    with fc3:
        codigo_query = st.text_input(
            "Código de monitoreo",
            key="bd_codigo",
            placeholder="Ej. E-1, RChil1…",
            help="Filtra por código del punto de monitoreo o de la muestra. "
                 "Coincidencia parcial, no distingue mayúsculas.",
        ).strip()
    with fc4:
        _hoy = date.today()
        _default_desde = _hoy.replace(year=_hoy.year - 1)
        fecha_inicio = st.date_input("Desde", value=_default_desde, key="bd_desde")
    with fc5:
        fecha_fin = st.date_input("Hasta", value=_hoy, key="bd_hasta")

    # Segunda fila: opciones de presentación (categorías + celdas vacías)
    oc1, oc2 = st.columns([3, 1])
    with oc1:
        _categorias_disponibles = list(get_cat_params().keys())
        categoria_filtro = st.multiselect(
            "Categorías a mostrar",
            _categorias_disponibles,
            default=_categorias_disponibles,
            key="bd_categorias",
        )
    with oc2:
        st.markdown("<div style='height:16px;'></div>", unsafe_allow_html=True)
        mostrar_vacios = st.checkbox("Mostrar celdas vacías", value=True, key="bd_vacios")

    # ── Cargar datos ────────────────────────────────────────────────────
    with st.spinner("Cargando base de datos..."):
        try:
            datos = get_datos_consolidados(
                campana_id=campana_id,
                punto_ids=punto_ids_filtro,
                fecha_inicio=str(fecha_inicio) if fecha_inicio else None,
                fecha_fin=str(fecha_fin) if fecha_fin else None,
            )
        except Exception as e:
            st.error(f"Error cargando datos consolidados: {type(e).__name__}: {e}")
            st.exception(e)
            st.stop()
        limites = get_limites_eca_todos()

    # Filtro por código de monitoreo (punto o muestra): coincidencia parcial
    # sobre los datos ya cargados. Se aplica antes de las métricas y del df para
    # que la tabla, el resumen y las descargas reflejen la búsqueda.
    if codigo_query:
        q = codigo_query.lower()
        datos = [
            d for d in datos
            if q in (d.get("punto_codigo") or "").lower()
            or q in (d.get("codigo_muestra") or "").lower()
        ]

    if not datos:
        st.info("No se encontraron resultados con los filtros seleccionados.")
        st.stop()

    # Si se alcanzó el techo de muestras, los datos podrían estar truncados:
    # avisar para que el usuario acote los filtros en lugar de mostrar un
    # subconjunto silenciosamente.
    if len(datos) >= LIMITE_MUESTRAS:
        st.warning(
            f"Se alcanzó el máximo de {LIMITE_MUESTRAS:,} muestras por consulta. "
            "Es posible que falten registros: acota el rango de fechas, la campaña "
            "o el lugar para ver el conjunto completo."
        )

    # ── Filtrar columnas por categoría (dinámico desde BD) ──────────────
    cat_params = get_cat_params()
    COLUMNAS_PARAMETROS = get_columnas_parametros()
    formato_codigo = _formato_por_codigo(cat_params)
    lcm_codigo = get_lcm_por_codigo()

    codigos_visibles = []
    for cat in categoria_filtro:
        codigos_visibles.extend(cat_params.get(cat, []))

    columnas_visibles = [(cod, label) for cod, label in COLUMNAS_PARAMETROS if cod in codigos_visibles]

    # ── Conteos para el resumen y la firma de caché ─────────────────────
    # No se muestran como tarjetas; n_muestras/n_puntos alimentan el resumen
    # compacto sobre la tabla y, junto con los filtros, discriminan la firma de
    # caché de las descargas. Se calculan sobre `datos` ya filtrado (incluido el
    # filtro por código de monitoreo).
    n_muestras = len(datos)
    n_puntos = len({d["punto_codigo"] for d in datos})

    # ── Atajos del flujo con la campaña seleccionada ────────────────────
    # Solo se muestran cuando hay una campaña elegida y el rol lo permite, para
    # no dejar una fila vacía con "Todas las campañas".
    _puede_ver_campana = rol_alcanza("administrador")
    _puede_ver_informe = rol_alcanza("visualizador")
    if campana_id and (_puede_ver_campana or _puede_ver_informe):
        nav1, nav2, _sp = st.columns([1.4, 1.4, 3.2])
        with nav1:
            if _puede_ver_campana and st.button(
                    "Ver campaña", key="bd_nav_campana",
                    icon=":material/event:", use_container_width=True):
                ir_a("campanas", campana_id=campana_id)
        with nav2:
            if _puede_ver_informe and st.button(
                    "Informe de campaña", key="bd_nav_informe",
                    icon=":material/description:", use_container_width=True):
                ir_a("informes", campana_id=campana_id)

    # ── Construir DataFrame para mostrar ────────────────────────────────
    # Orden cronológico ascendente. Se agrupa PRIMERO por campaña (fecha_inicio
    # + código) para que los separadores amarillos no se repitan; dentro de cada
    # campaña las muestras van por fecha y luego por HORA de recolección (las más
    # tempranas primero), con el código como desempate. Así, al editar la hora de
    # una muestra, esta se reubica cronológicamente de forma automática.
    datos = sorted(
        datos,
        key=lambda d: (
            d.get("campana_fecha_inicio") or d.get("fecha") or "",
            d.get("campana_codigo") or "",
            d.get("campana_id") or "",
            d.get("fecha") or "",
            _hora_a_minutos(d.get("hora")),
            d.get("punto_codigo") or "",
            d.get("codigo_muestra") or "",
        ),
    )

    df_rows = []
    for d in datos:
        row = {
            "Fecha": d["fecha"],
            "Hora": d.get("hora", ""),
            "Código Punto": d["punto_codigo"],
            "Punto": d["punto_nombre"],
            "Código Muestra": d.get("codigo_muestra", ""),
            "Código Lab.": d.get("codigo_laboratorio") or "",
            "Profundidad (m)": d.get("profundidad"),
            "Cuenca": d["cuenca"],
            "Tipo": (d["tipo"] or "").capitalize(),
            "ECA": d["eca_codigo"],
        }
        for cod, label in columnas_visibles:
            row[label] = d.get(cod)
        df_rows.append(row)

    # Ocultar columna "Código Lab." si ninguna muestra tiene valor asignado.
    # Reaparece automáticamente en cuanto se cargue un código en Recepción.
    if all(not r.get("Código Lab.") for r in df_rows):
        for r in df_rows:
            r.pop("Código Lab.", None)

    # Si ninguna muestra tiene profundidad registrada, también se oculta
    # para mantener la vista limpia en los puntos superficiales.
    if all(r.get("Profundidad (m)") in (None, "") for r in df_rows):
        for r in df_rows:
            r.pop("Profundidad (m)", None)

    df = pd.DataFrame(df_rows)

    # ── Tabs: Vista y Edición ───────────────────────────────────────────
    if es_admin:
        tab_vista, tab_edicion = st.tabs(["Vista consulta", "Edición de datos"])
    else:
        tab_vista = st.container()
        tab_edicion = None

    # ── Tab Vista ───────────────────────────────────────────────────────
    with tab_vista:
        editar_inline = False
        if es_admin:
            editar_inline = st.toggle(
                "Editar en la tabla",
                key="bd_editar_inline",
                help="Edita directamente los valores y metadatos de los registros "
                     "visibles. Los cambios se guardan en Resultados e Informes.",
            )

        st.markdown(f"**{n_muestras} registros** · {n_puntos} puntos · "
                    f"Las celdas en **rojo** exceden su ECA respectivo")

        # Paginación del render: con muchos registros, construir la tabla HTML
        # completa genera un DOM enorme y vuelve lenta la página. Mostramos solo
        # una página de filas; las métricas (arriba) y las descargas (abajo)
        # siguen cubriendo el conjunto completo. Con pocos registros no se pagina.
        pagina = 1
        total_filas = len(datos)
        if total_filas > _FILAS_POR_PAGINA:
            n_paginas = (total_filas + _FILAS_POR_PAGINA - 1) // _FILAS_POR_PAGINA
            pcol1, pcol2 = st.columns([1, 3])
            with pcol1:
                pagina = int(st.number_input(
                    f"Página (de {n_paginas})",
                    min_value=1, max_value=n_paginas, value=1, step=1,
                    key="bd_pagina",
                ))
            inicio = (pagina - 1) * _FILAS_POR_PAGINA
            fin = min(inicio + _FILAS_POR_PAGINA, total_filas)
            with pcol2:
                st.caption(
                    f"Mostrando registros {inicio + 1:,}–{fin:,} de {total_filas:,}. "
                    "Las descargas incluyen el conjunto completo."
                )
            datos_vista = datos[inicio:fin]
            df_vista = df.iloc[inicio:fin]
        else:
            datos_vista = datos
            df_vista = df

        if editar_inline:
            st.caption("✏️ Editando los registros de esta página. Cambia de página "
                       "o ajusta los filtros para editar otras muestras. Desactiva "
                       "el interruptor para volver a la vista con colores.")
            if not columnas_visibles:
                st.info("No hay parámetros seleccionados (revisa **Categorías a "
                        "mostrar** arriba). Puedes editar los metadatos de la muestra.")
            _render_grid_editor(
                datos_subset=datos_vista,
                columnas_visibles=columnas_visibles,
                formato_codigo=formato_codigo,
                limites=limites,
                param_map=get_parametros_map(),
                key_prefix=f"bd_inline_p{pagina}",
            )
        else:
            html_table = _render_tabla_por_campana(
                df=df_vista,
                datos=datos_vista,
                columnas_visibles=columnas_visibles,
                formato_codigo=formato_codigo,
                limites=limites,
                lcm_codigo=lcm_codigo,
            )
            st.markdown(html_table, unsafe_allow_html=True)

        # ── Descargas (siempre el conjunto completo filtrado) ────────────
        sig = (f"{campana_id}|{punto_ids_filtro}|{fecha_inicio}|{fecha_fin}|"
               f"{tuple(categoria_filtro)}|{codigo_query}|{n_muestras}|"
               f"v{st.session_state.get('bd_data_version', 0)}")
        dl1, dl2, _dsp = st.columns([1.1, 1.1, 2.8])
        with dl1:
            try:
                xlsx_bytes = _xlsx_cacheado(
                    sig, df, datos, columnas_visibles, formato_codigo, limites, lcm_codigo,
                )
                st.download_button(
                    "Descargar Excel",
                    xlsx_bytes,
                    f"base_datos_lvca_{date.today()}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="bd_download_xlsx",
                    icon=":material/download:",
                    use_container_width=True,
                )
            except Exception as e:
                st.caption(f"No se pudo generar el Excel: {type(e).__name__}: {e}")
        with dl2:
            st.download_button(
                "Descargar CSV",
                _df_a_csv(df),
                f"base_datos_lvca_{date.today()}.csv",
                "text/csv",
                key="bd_download",
                icon=":material/table_view:",
                use_container_width=True,
            )

    # ── Tab Edición (solo admin): cuadrícula por campaña ────────────────
    if tab_edicion is not None:
        with tab_edicion:
            st.markdown(
                "Elige una **campaña** para editar todas sus muestras en una "
                "cuadrícula tipo Excel. Los cambios en valores y metadatos se "
                "reflejan automáticamente en **Resultados** e **Informes**."
            )

            campanas_ed = get_campanas()
            if not campanas_ed:
                st.info("No hay campañas disponibles para editar.")
            else:
                op_ed = {f"{c['codigo']} — {c['nombre']}": c["id"] for c in campanas_ed}
                claves_ed = list(op_ed.keys())

                # Preseleccionar la campaña elegida en el filtro superior, si la hay.
                idx_def = 0
                if campana_id:
                    for _pos, _k in enumerate(claves_ed):
                        if op_ed[_k] == campana_id:
                            idx_def = _pos
                            break

                sel_ed = st.selectbox(
                    "Campaña a editar", claves_ed, index=idx_def, key="bd_ed_camp",
                    help="Escribe para buscar. Se cargarán todas las muestras de la campaña.",
                )
                camp_ed_id = op_ed[sel_ed]

                with st.spinner("Cargando muestras de la campaña..."):
                    datos_ed = get_datos_consolidados(campana_id=camp_ed_id)

                datos_ed = sorted(
                    datos_ed,
                    key=lambda d: (d.get("fecha") or "", _hora_a_minutos(d.get("hora")),
                                   d.get("punto_codigo") or "", d.get("codigo_muestra") or ""),
                )

                if not datos_ed:
                    st.info("Esta campaña no tiene muestras registradas.")
                elif not columnas_visibles:
                    st.warning("Selecciona al menos una categoría en los filtros de "
                               "arriba para ver los parámetros editables.")
                else:
                    st.caption(
                        f"{len(datos_ed)} muestra(s) · {len(columnas_visibles)} "
                        "parámetro(s) según las categorías seleccionadas arriba."
                    )
                    _render_grid_editor(
                        datos_subset=datos_ed,
                        columnas_visibles=columnas_visibles,
                        formato_codigo=formato_codigo,
                        limites=limites,
                        param_map=get_parametros_map(),
                        key_prefix=f"bd_ed_{camp_ed_id}",
                    )


main()
