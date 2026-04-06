"""
database/seeds/seed_puntos_2026.py
Carga los 27 puntos de monitoreo desde BASE DE DATOS 2026 - LVCA.xlsx
UPSERT por codigo. Desactiva puntos no incluidos.

Requiere haber ejecutado migrations/003_puntos_completo.sql en Supabase primero.

Ejecutar:
    cd c:/proyectos/lvca_agua && python -m database.seeds.seed_puntos_2026
"""

import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "../..")))

import openpyxl
from database.client import get_admin_client

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "../../BASE DE DATOS 2026 - LVCA.xlsx")

# Mapeo tipo de monitoreo → valor en BD
TIPO_MAP = {
    "agua de embalse":      "embalse",
    "agua de rio":          "rio",
    "agua de bocatoma":     "bocatoma",
    "agua de desarenador":  "desarenador",
}

# Normalización de cuenca
CUENCA_MAP = {
    "132 quilca - chili - vitor":  "Quilca-Chili-Vitor",
    "132 quilca - vitor - chili":  "Quilca-Chili-Vitor",
    "134 colca - camana":          "Colca-Camaná",
}


def _cuenca(raw: str) -> str:
    return CUENCA_MAP.get((raw or "").strip().lower(), (raw or "").strip())


def _zona(raw: str) -> str:
    return (raw or "").replace(" ", "").upper()  # "19 L" → "19L"


def _tipo(raw: str) -> str:
    return TIPO_MAP.get((raw or "").strip().lower(), "embalse")


def run():
    db = get_admin_client()

    # ── Obtener IDs de ECAs ────────────────────────────────────────────────
    ecas_raw = db.table("ecas").select("id, codigo").execute().data or []
    eca_map = {e["codigo"]: e["id"] for e in ecas_raw}
    print(f"ECAs disponibles: {list(eca_map.keys())}")

    # ── Leer Excel ─────────────────────────────────────────────────────────
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["BASE INICIAL"]
    headers = [c.value for c in ws[1]]
    print(f"Columnas Excel: {headers}")

    h = {v: i for i, v in enumerate(headers)}

    puntos_excel = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo = row[h["CODIGO_LUGAR"]]
        if not codigo:
            continue

        eca_raw = str(row[h["ECA"]] or "").strip()
        eca_id  = eca_map.get(eca_raw)
        if not eca_id:
            print(f"  AVISO: ECA '{eca_raw}' no encontrado para {codigo}")

        zona_raw = str(row[h["ZONA_UTM"]] or "").strip()

        punto = {
            "codigo":           str(codigo).strip(),
            "nombre":           str(row[h["LUGAR_DE_MUESTREO"]] or "").strip(),
            "descripcion":      str(row[h["REPRESENTATIVIDAD"]] or "").strip() or None,
            "cuenca":           _cuenca(str(row[h["CUENCA"]] or "")),
            "subcuenca":        str(row[h["SISTEMA_HIDRICO"]] or "").strip() or None,
            "tipo":             _tipo(str(row[h["MONITOREO"]] or "")),
            "utm_zona":         _zona(zona_raw),
            "utm_este":         row[h["ESTE (m)"]],
            "utm_norte":        row[h["NORTE (m)"]],
            "altitud_msnm":     row[h["ALTITUD (m s.n.m.)"]],
            "departamento":     str(row[h["DEPARTAMENTO"]] or "").strip().upper() or None,
            "provincia":        str(row[h["PROVINCIA"]] or "").strip().upper() or None,
            "distrito":         str(row[h["DISTRITO"]] or "").strip().upper() or None,
            "accesibilidad":    str(row[h["ACCESIBILIDAD"]] or "").strip() or None,
            "representatividad":str(row[h["REPRESENTATIVIDAD"]] or "").strip() or None,
            "finalidad":        str(row[h["FINALIDAD"]] or "").strip() or None,
            "sistema_hidrico":  str(row[h["SISTEMA_HIDRICO"]] or "").strip() or None,
            "lugar_muestreo":   str(row[h["LUGAR_DE_MUESTREO"]] or "").strip() or None,
            "eca_id":           eca_id,
            "activo":           True,
            "entidad_responsable": "AUTODEMA",
        }
        puntos_excel.append(punto)

    print(f"\nPuntos a cargar desde Excel: {len(puntos_excel)}")

    # ── Upsert puntos ──────────────────────────────────────────────────────
    codigos_excel = {p["codigo"] for p in puntos_excel}
    ok = 0
    errores = []

    for p in puntos_excel:
        try:
            db.table("puntos_muestreo").upsert(p, on_conflict="codigo").execute()
            print(f"  OK: {p['codigo']} — {p['nombre']}")
            ok += 1
        except Exception as e:
            print(f"  ERROR: {p['codigo']} — {e}")
            errores.append(p["codigo"])

    # ── Desactivar puntos que ya no están en el Excel ──────────────────────
    todos = db.table("puntos_muestreo").select("id, codigo, activo").execute().data or []
    desactivados = 0
    for p in todos:
        if p["codigo"] not in codigos_excel and p.get("activo"):
            db.table("puntos_muestreo").update({"activo": False}).eq("id", p["id"]).execute()
            print(f"  DESACTIVADO: {p['codigo']}")
            desactivados += 1

    print(f"\n{'='*50}")
    print(f"Resultado: {ok} OK | {len(errores)} errores | {desactivados} desactivados")
    if errores:
        print(f"Errores en: {errores}")


if __name__ == "__main__":
    run()
